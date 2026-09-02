from __future__ import annotations

import io
import os
import re
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader

from warehouse_structure import ESTRUTURA_CD, gerar_todos_casulos


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("OUTLOG_DATA_DIR", str(BASE_DIR / "data"))).resolve() / "controle_logistica.db"


def db_connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    return con


def now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def norm(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", text)


def as_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value or default).replace(".", "").replace(",", ".")))
    except (TypeError, ValueError):
        return default


def read_pdf_text(raw: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(raw))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:
        raise HTTPException(400, f"Não foi possível ler o PDF: {exc}")


def parse_pdf_items(raw: bytes) -> tuple[str, list[dict[str, Any]]]:
    """Extrai itens de relatórios tabulares sem depender de um leiaute único."""
    text = read_pdf_text(raw)
    document_match = re.search(r"(?:romaneio|documento|nota|pedido)\s*(?:n[ºo°.]*)?\s*[:#-]?\s*([A-Z0-9./-]{3,})", text, re.I)
    document_no = document_match.group(1) if document_match else ""
    parsed: dict[str, dict[str, Any]] = {}
    for raw_line in text.splitlines():
        line = " ".join(raw_line.split())
        barcode_match = re.search(r"\b(\d{8,14})\b", line)
        if not barcode_match:
            continue
        barcode = barcode_match.group(1)
        remainder = (line[:barcode_match.start()] + " " + line[barcode_match.end():]).strip()
        numbers = [as_int(v) for v in re.findall(r"(?<![A-Za-z])\d{1,6}(?![A-Za-z])", remainder)]
        quantity = next((n for n in reversed(numbers) if n > 0), 1)
        description = re.sub(r"\b\d{1,6}\b", " ", remainder)
        description = " ".join(description.split()).strip(" -|;") or "Item importado do PDF"
        if barcode in parsed:
            parsed[barcode]["quantity"] += quantity
        else:
            parsed[barcode] = {"barcode": barcode, "description": description[:180], "quantity": quantity}
    if not parsed:
        raise HTTPException(400, "Nenhum item com código de barras foi reconhecido. Confira se o PDF possui texto pesquisável.")
    return document_no, list(parsed.values())


def require_user(con: sqlite3.Connection, user_id: int, roles: set[str] | None = None) -> sqlite3.Row:
    row = con.execute("SELECT id,username,name,role FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        raise HTTPException(401, "Usuário inválido.")
    if roles and row["role"] not in roles:
        raise HTTPException(403, "Seu perfil não permite esta operação.")
    return row


def init_unified_db() -> None:
    con = db_connect()
    con.executescript(
        """
        CREATE TABLE IF NOT EXISTS warehouse_zones(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          code TEXT NOT NULL UNIQUE,
          name TEXT NOT NULL,
          gender TEXT,
          capacity INTEGER NOT NULL DEFAULT 0,
          active INTEGER NOT NULL DEFAULT 1,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS warehouse_locations(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          zone_id INTEGER NOT NULL REFERENCES warehouse_zones(id),
          address TEXT NOT NULL UNIQUE,
          aisle TEXT NOT NULL,
          side TEXT,
          column_no INTEGER,
          level_no TEXT,
          structure_type TEXT,
          category TEXT,
          capacity INTEGER NOT NULL DEFAULT 0,
          occupied_qty INTEGER NOT NULL DEFAULT 0,
          status TEXT NOT NULL DEFAULT 'DISPONIVEL',
          updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_entries(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          card_id INTEGER REFERENCES cards(id),
          item_id INTEGER REFERENCES items(id),
          location_id INTEGER NOT NULL REFERENCES warehouse_locations(id),
          quantity INTEGER NOT NULL CHECK(quantity > 0),
          brand TEXT,
          category TEXT,
          lot TEXT,
          created_by INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_report_imports(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          filename TEXT NOT NULL,
          total_qty INTEGER NOT NULL DEFAULT 0,
          imported_by INTEGER REFERENCES users(id),
          imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_group_snapshots(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          import_id INTEGER NOT NULL REFERENCES stock_report_imports(id) ON DELETE CASCADE,
          group_name TEXT NOT NULL,
          gender TEXT NOT NULL DEFAULT 'OUTROS',
          quantity INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS sgo_imports(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          filename TEXT NOT NULL,
          row_count INTEGER NOT NULL DEFAULT 0,
          imported_by INTEGER REFERENCES users(id),
          imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sgo_entries(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          import_id INTEGER REFERENCES sgo_imports(id),
          purchase_id TEXT,
          group_name TEXT,
          description TEXT,
          brand TEXT,
          quantity INTEGER NOT NULL DEFAULT 0,
          status TEXT NOT NULL DEFAULT 'EM_TRANSITO',
          forecast_date TEXT,
          origin TEXT,
          mapped_type TEXT,
          updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS operational_tasks(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          title TEXT NOT NULL,
          description TEXT,
          sector TEXT NOT NULL,
          status TEXT NOT NULL DEFAULT 'PENDENTE',
          priority TEXT NOT NULL DEFAULT 'NORMAL',
          responsible_id INTEGER REFERENCES users(id),
          card_id INTEGER REFERENCES cards(id),
          sgo_entry_id INTEGER REFERENCES sgo_entries(id),
          expected_qty INTEGER NOT NULL DEFAULT 0,
          created_by INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL,
          completed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS shipments(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          document_no TEXT NOT NULL,
          destination TEXT,
          status TEXT NOT NULL DEFAULT 'PREPARANDO',
          total_qty INTEGER NOT NULL DEFAULT 0,
          created_by INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL,
          completed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS shipment_items(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          shipment_id INTEGER NOT NULL REFERENCES shipments(id) ON DELETE CASCADE,
          stock_entry_id INTEGER REFERENCES stock_entries(id),
          barcode TEXT,
          description TEXT,
          quantity INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS returns(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          document_no TEXT NOT NULL,
          store TEXT,
          customer TEXT,
          status TEXT NOT NULL DEFAULT 'RECEBIDA',
          total_store INTEGER NOT NULL DEFAULT 0,
          total_cd INTEGER NOT NULL DEFAULT 0,
          total_anapolis INTEGER NOT NULL DEFAULT 0,
          difference INTEGER NOT NULL DEFAULT 0,
          decision TEXT,
          notes TEXT,
          created_by INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL,
          completed_at TEXT
        );
        CREATE TABLE IF NOT EXISTS return_items(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          return_id INTEGER NOT NULL REFERENCES returns(id) ON DELETE CASCADE,
          barcode TEXT,
          reference TEXT,
          description TEXT,
          size TEXT,
          qty_store INTEGER NOT NULL DEFAULT 0,
          qty_cd INTEGER NOT NULL DEFAULT 0,
          qty_anapolis INTEGER NOT NULL DEFAULT 0,
          difference INTEGER NOT NULL DEFAULT 0,
          condition_status TEXT NOT NULL DEFAULT 'PENDENTE',
          destination TEXT,
          notes TEXT
        );
        CREATE TABLE IF NOT EXISTS unified_movements(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          domain TEXT NOT NULL,
          entity_id INTEGER,
          event_type TEXT NOT NULL,
          description TEXT NOT NULL,
          user_id INTEGER REFERENCES users(id),
          created_at TEXT NOT NULL
        );
        """
    )
    shipment_columns = {row["name"] for row in con.execute("PRAGMA table_info(shipments)").fetchall()}
    for name, definition in {
        "carrier": "TEXT",
        "vehicle_plate": "TEXT",
        "driver_name": "TEXT",
        "scheduled_at": "TEXT",
        "volume_count": "INTEGER NOT NULL DEFAULT 0",
        "notes": "TEXT",
        "updated_at": "TEXT",
    }.items():
        if name not in shipment_columns:
            con.execute(f"ALTER TABLE shipments ADD COLUMN {name} {definition}")
    shipment_item_columns = {row["name"] for row in con.execute("PRAGMA table_info(shipment_items)").fetchall()}
    for name, definition in {
        "checked_qty": "INTEGER NOT NULL DEFAULT 0",
        "notes": "TEXT",
    }.items():
        if name not in shipment_item_columns:
            con.execute(f"ALTER TABLE shipment_items ADD COLUMN {name} {definition}")
    if not con.execute("SELECT 1 FROM warehouse_zones LIMIT 1").fetchone():
        zones = [
            ("A", "Zona A — Giro alto", "Misto", 5000),
            ("B", "Zona B — Grade", "Grade", 4500),
            ("C", "Zona C — Saldo", "Saldo", 3500),
            ("D", "Zona D — Reserva", "Misto", 2500),
        ]
        con.executemany(
            "INSERT INTO warehouse_zones(code,name,gender,capacity,created_at) VALUES(?,?,?,?,?)",
            [(code, name, gender, capacity, now()) for code, name, gender, capacity in zones],
        )
        zone_ids = {r["code"]: r["id"] for r in con.execute("SELECT id,code FROM warehouse_zones")}
        locations = []
        for zone in "ABCD":
            for column in range(1, 9):
                for level in ("A", "B", "C"):
                    address = f"{zone}-{column:03d}-{level}"
                    locations.append((zone_ids[zone], address, zone, "UNICO", column, level, "CASULO", 100, now()))
        con.executemany(
            """INSERT INTO warehouse_locations(zone_id,address,aisle,side,column_no,level_no,structure_type,capacity,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            locations,
        )
    # Integração segura da estrutura física real do CD. Os endereços antigos
    # são preservados no banco (inclusive seus lançamentos de estoque), mas
    # ficam inativos quando a estrutura real está pronta. Nenhum registro é
    # apagado durante a migração.
    casulos = gerar_todos_casulos()
    real_location_count = con.execute(
        """SELECT COUNT(*) n FROM warehouse_locations l
           JOIN warehouse_zones z ON z.id=l.zone_id WHERE z.code LIKE 'Rua %'"""
    ).fetchone()["n"]
    if real_location_count < len(casulos):
        ruas_presentes = sorted({c["rua"] for c in casulos}, key=lambda r: int(r.split()[1]))
        zone_capacity: dict[str, int] = {}
        zone_gender: dict[str, str] = {}
        for casulo in casulos:
            zone_capacity[casulo["rua"]] = zone_capacity.get(casulo["rua"], 0) + int(casulo["capacidade"])
            zone_gender[casulo["rua"]] = casulo["genero"]
        con.executemany(
            """INSERT OR IGNORE INTO warehouse_zones(code,name,gender,capacity,created_at)
               VALUES(?,?,?,?,?)""",
            [(rua, rua, zone_gender[rua], zone_capacity[rua], now()) for rua in ruas_presentes],
        )
        zone_ids = {r["code"]: r["id"] for r in con.execute(
            "SELECT id,code FROM warehouse_zones WHERE code LIKE 'Rua %'"
        )}
        con.executemany(
            """INSERT OR IGNORE INTO warehouse_locations(
               zone_id,address,aisle,side,column_no,level_no,structure_type,capacity,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            [(
                zone_ids[c["rua"]], c["address"], c["rua"], c["lado"], c["coluna"],
                c["nivel"], c["tipo_estrutural"], c["capacidade"], now(),
            ) for c in casulos],
        )
        con.execute("UPDATE warehouse_zones SET active=0 WHERE code NOT LIKE 'Rua %'")
        con.execute("UPDATE warehouse_zones SET active=1 WHERE code LIKE 'Rua %'")
        add_movement(
            con, "ESTOQUE", None, "SEED_ESTRUTURA_REAL_SEGURA",
            f"Estrutura real carregada com {len(casulos)} casulos; endereços legados preservados e inativados.",
            None,
        )
    con.commit()
    con.close()


def add_movement(con: sqlite3.Connection, domain: str, entity_id: int | None, event: str, description: str, user_id: int | None) -> None:
    con.execute(
        "INSERT INTO unified_movements(domain,entity_id,event_type,description,user_id,created_at) VALUES(?,?,?,?,?,?)",
        (domain, entity_id, event, description, user_id, now()),
    )


def register_unified_routes(app) -> None:
    @app.get("/api/unified/overview")
    def overview():
        con = db_connect()
        sectors = {r["current_sector"]: r["n"] for r in con.execute("SELECT current_sector,COUNT(*) n FROM cards GROUP BY current_sector")}
        warehouse = con.execute(
            """SELECT COALESCE(SUM(l.capacity),0) capacity,COALESCE(SUM(l.occupied_qty),0) occupied,COUNT(*) locations
               FROM warehouse_locations l JOIN warehouse_zones z ON z.id=l.zone_id WHERE z.active=1"""
        ).fetchone()
        task_counts = {r["status"]: r["n"] for r in con.execute("SELECT status,COUNT(*) n FROM operational_tasks GROUP BY status")}
        return_counts = {r["status"]: r["n"] for r in con.execute("SELECT status,COUNT(*) n FROM returns GROUP BY status")}
        sgo_counts = {r["status"]: r["n"] for r in con.execute("SELECT status,COUNT(*) n FROM sgo_entries GROUP BY status")}
        recent = con.execute(
            """SELECT domain,event_type,description,created_at FROM unified_movements
               ORDER BY id DESC LIMIT 12"""
        ).fetchall()
        capacity = int(warehouse["capacity"] or 0)
        occupied = int(warehouse["occupied"] or 0)
        con.close()
        return {
            "sectors": sectors,
            "warehouse": {"capacity": capacity, "occupied": occupied, "percentage": round(occupied * 100 / capacity, 1) if capacity else 0, "locations": warehouse["locations"]},
            "tasks": task_counts,
            "returns": return_counts,
            "sgo": sgo_counts,
            "recent_movements": [dict(r) for r in recent],
        }

    @app.get("/api/unified/warehouse")
    def warehouse(search: str = "", zone: str = "", status: str = ""):
        con = db_connect()
        sql = """SELECT l.*,z.code zone_code,z.name zone_name,
                 CASE WHEN l.capacity>0 THEN ROUND(l.occupied_qty*100.0/l.capacity,1) ELSE 0 END occupancy
                 FROM warehouse_locations l JOIN warehouse_zones z ON z.id=l.zone_id WHERE z.active=1"""
        args: list[Any] = []
        if search:
            sql += " AND (l.address LIKE ? OR l.category LIKE ? OR l.structure_type LIKE ?)"
            q = f"%{search}%"; args += [q, q, q]
        if zone:
            sql += " AND z.code=?"; args.append(zone)
        if status:
            sql += " AND l.status=?"; args.append(status)
        sql += " ORDER BY z.code,l.column_no,l.level_no"
        locations = [dict(r) for r in con.execute(sql, args).fetchall()]
        zones = [dict(r) for r in con.execute(
            """SELECT z.*,COALESCE(SUM(l.occupied_qty),0) occupied,
               CASE WHEN z.capacity>0 THEN ROUND(COALESCE(SUM(l.occupied_qty),0)*100.0/z.capacity,1) ELSE 0 END occupancy
               FROM warehouse_zones z LEFT JOIN warehouse_locations l ON l.zone_id=z.id
               WHERE z.active=1 GROUP BY z.id ORDER BY z.code"""
        ).fetchall()]
        con.close()
        return {"zones": zones, "locations": locations}

    @app.post("/api/unified/warehouse/locations/{location_id}")
    async def update_location(location_id: int, request: Request):
        data = await request.json(); user_id = as_int(data.get("user_id"))
        con = db_connect(); require_user(con, user_id, {"admin", "supervisor", "estocagem"})
        location = con.execute("SELECT * FROM warehouse_locations WHERE id=?", (location_id,)).fetchone()
        if not location:
            con.close(); raise HTTPException(404, "Endereço não encontrado.")
        capacity = max(0, as_int(data.get("capacity"), location["capacity"]))
        category = str(data.get("category", location["category"] or "")).strip()
        status = str(data.get("status", location["status"])).upper()
        con.execute("UPDATE warehouse_locations SET capacity=?,category=?,status=?,updated_at=? WHERE id=?", (capacity, category, status, now(), location_id))
        add_movement(con, "ESTOQUE", location_id, "ENDERECO_ATUALIZADO", f"Endereço {location['address']} atualizado.", user_id)
        con.commit(); con.close(); return {"ok": True}

    @app.post("/api/unified/warehouse/store")
    async def store(request: Request):
        data = await request.json(); user_id = as_int(data.get("user_id")); qty = as_int(data.get("quantity"))
        con = db_connect(); require_user(con, user_id, {"admin", "supervisor", "estocagem"})
        location = con.execute("SELECT * FROM warehouse_locations WHERE id=?", (as_int(data.get("location_id")),)).fetchone()
        if not location or qty <= 0:
            con.close(); raise HTTPException(400, "Informe endereço e quantidade válidos.")
        if int(location["occupied_qty"]) + qty > int(location["capacity"]):
            con.close(); raise HTTPException(400, "A quantidade ultrapassa a capacidade do endereço.")
        card_id = as_int(data.get("card_id")) or None; item_id = as_int(data.get("item_id")) or None
        cur = con.execute(
            """INSERT INTO stock_entries(card_id,item_id,location_id,quantity,brand,category,lot,created_by,created_at)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            (card_id, item_id, location["id"], qty, str(data.get("brand", "")), str(data.get("category", "")), str(data.get("lot", "")), user_id, now()),
        )
        con.execute("UPDATE warehouse_locations SET occupied_qty=occupied_qty+?,status=CASE WHEN occupied_qty+?>=capacity THEN 'LOTADO' ELSE 'OCUPADO' END,updated_at=? WHERE id=?", (qty, qty, now(), location["id"]))
        if card_id:
            con.execute("UPDATE cards SET casulo_current=?,current_sector='ESTOCAGEM',status='FINALIZADO',updated_at=? WHERE id=?", (location["address"], now(), card_id))
        add_movement(con, "ESTOQUE", cur.lastrowid, "ENTRADA_ESTOQUE", f"{qty} peça(s) armazenadas em {location['address']}.", user_id)
        con.commit(); con.close(); return {"ok": True, "stock_entry_id": cur.lastrowid}

    @app.get("/api/unified/warehouse/analytics")
    def warehouse_analytics():
        con=db_connect()
        structures=[dict(r) for r in con.execute("""SELECT COALESCE(structure_type,'NÃO DEFINIDO') label,COUNT(*) locations,
            COALESCE(SUM(capacity),0) capacity,COALESCE(SUM(occupied_qty),0) occupied FROM warehouse_locations GROUP BY structure_type ORDER BY locations DESC""").fetchall()]
        categories=[dict(r) for r in con.execute("""SELECT COALESCE(NULLIF(category,''),'SEM CATEGORIA') label,
            COALESCE(SUM(quantity),0) quantity FROM stock_entries GROUP BY category ORDER BY quantity DESC LIMIT 20""").fetchall()]
        brands=[dict(r) for r in con.execute("""SELECT COALESCE(NULLIF(brand,''),'SEM MARCA') label,
            COALESCE(SUM(quantity),0) quantity FROM stock_entries GROUP BY brand ORDER BY quantity DESC LIMIT 20""").fetchall()]
        latest=con.execute("SELECT * FROM stock_report_imports ORDER BY id DESC LIMIT 1").fetchone()
        groups=[]
        if latest:
            groups=[dict(r) for r in con.execute("SELECT * FROM stock_group_snapshots WHERE import_id=? ORDER BY quantity DESC",(latest["id"],)).fetchall()]
        con.close();return {"structures":structures,"categories":categories,"brands":brands,"latest_report":dict(latest) if latest else None,"groups":groups}

    @app.post("/api/unified/warehouse/import-group-report")
    async def import_group_report(user_id:int,file:UploadFile=File(...)):
        text=read_pdf_text(await file.read());groups={}
        for raw_line in text.splitlines():
            line=" ".join(raw_line.split()).strip()
            match=re.match(r"^(.{3,}?)\s+([\d.]{1,12})$",line)
            if not match or re.search(r"total|quantidade|estoque do grupo",match.group(1),re.I):continue
            name=match.group(1).strip(" -:|")[:120];qty=as_int(match.group(2))
            if name and qty>0:groups[name]=groups.get(name,0)+qty
        if not groups:raise HTTPException(400,"Nenhum grupo foi reconhecido nesse PDF. Confira se é o Resumo de Estoque do Grupo com texto pesquisável.")
        con=db_connect();require_user(con,user_id,{"admin","supervisor","estocagem"})
        total=sum(groups.values());cur=con.execute("INSERT INTO stock_report_imports(filename,total_qty,imported_by,imported_at) VALUES(?,?,?,?)",(file.filename or "estoque-grupos.pdf",total,user_id,now()))
        for name,qty in groups.items():
            key=norm(name);gender="FEMININO" if "femin" in key else "MASCULINO" if "masc" in key else "OUTROS"
            con.execute("INSERT INTO stock_group_snapshots(import_id,group_name,gender,quantity) VALUES(?,?,?,?)",(cur.lastrowid,name,gender,qty))
        add_movement(con,"ESTOQUE",cur.lastrowid,"RELATORIO_GRUPOS",f"Relatório de estoque importado: {len(groups)} grupos e {total} peças.",user_id)
        con.commit();con.close();return {"ok":True,"groups":len(groups),"total_qty":total}

    @app.get("/api/unified/sgo")
    def list_sgo(status: str = "", search: str = ""):
        con = db_connect(); sql = "SELECT * FROM sgo_entries WHERE 1=1"; args: list[Any] = []
        if status: sql += " AND status=?"; args.append(status)
        if search:
            q = f"%{search}%"; sql += " AND (purchase_id LIKE ? OR group_name LIKE ? OR description LIKE ? OR brand LIKE ?)"; args += [q, q, q, q]
        sql += " ORDER BY COALESCE(forecast_date,'9999-12-31'),id DESC"
        rows = [dict(r) for r in con.execute(sql, args).fetchall()]
        con.close(); return rows

    @app.post("/api/unified/sgo/import")
    async def import_sgo(user_id: int, file: UploadFile = File(...)):
        raw = await file.read()
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(400, f"Não foi possível abrir o relatório: {exc}")
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: raise HTTPException(400, "Relatório vazio.")
        headers = {norm(v): i for i, v in enumerate(rows[0]) if v is not None}
        aliases = {
            "purchase_id": ("idcompra", "compra", "pedido", "lote"),
            "group": ("grupo", "tipoproduto", "categoria"),
            "description": ("descricao", "produto", "mercadoria"),
            "brand": ("marca",),
            "qty": ("quantidade", "qtd", "qtdesperada"),
            "status": ("status", "statussgo", "statuskanban"),
            "forecast": ("previsao", "previsaochegada", "dataprevisao"),
            "origin": ("origem", "fornecedor"),
        }
        def cell(row, key, default=""):
            idx = next((headers[a] for a in aliases[key] if a in headers), None)
            return default if idx is None or idx >= len(row) else row[idx]
        con = db_connect(); require_user(con, user_id, {"admin", "supervisor", "recebimento", "planejamento"})
        cur = con.execute("INSERT INTO sgo_imports(filename,row_count,imported_by,imported_at) VALUES(?,?,?,?)", (file.filename or "sgo.xlsx", 0, user_id, now()))
        import_id = cur.lastrowid; inserted = 0
        for row in rows[1:]:
            if not any(v not in (None, "") for v in row): continue
            status_raw = str(cell(row, "status", "EM_TRANSITO")).upper()
            status_key = "QUALIDADE" if "QUAL" in status_raw else "PROCESSAMENTO" if "PROCESS" in status_raw else "ESTOCAGEM" if "ESTOC" in status_raw else "CONCLUIDO" if "CONCLU" in status_raw else "EM_TRANSITO"
            con.execute(
                """INSERT INTO sgo_entries(import_id,purchase_id,group_name,description,brand,quantity,status,forecast_date,origin,mapped_type,updated_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (import_id, str(cell(row,"purchase_id","")), str(cell(row,"group","")), str(cell(row,"description","")), str(cell(row,"brand","")), as_int(cell(row,"qty",0)), status_key, str(cell(row,"forecast","")), str(cell(row,"origin","")), "", now()),
            ); inserted += 1
        con.execute("UPDATE sgo_imports SET row_count=? WHERE id=?", (inserted, import_id))
        add_movement(con, "SGO", import_id, "IMPORTACAO", f"Relatório SGO importado com {inserted} linha(s).", user_id)
        con.commit(); con.close(); return {"ok": True, "rows": inserted}

    @app.post("/api/unified/sgo/{entry_id}/task")
    async def sgo_to_task(entry_id: int, request: Request):
        data = await request.json(); user_id = as_int(data.get("user_id"))
        con = db_connect(); require_user(con, user_id, {"admin", "supervisor", "recebimento", "processamento", "qualidade", "estocagem", "planejamento"})
        entry = con.execute("SELECT * FROM sgo_entries WHERE id=?", (entry_id,)).fetchone()
        if not entry: con.close(); raise HTTPException(404, "Entrada SGO não encontrada.")
        sector = str(data.get("sector") or entry["status"] or "RECEBIMENTO").upper()
        cur = con.execute(
            """INSERT INTO operational_tasks(title,description,sector,status,priority,responsible_id,sgo_entry_id,expected_qty,created_by,created_at)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (f"SGO {entry['purchase_id'] or entry['group_name']}", entry["description"], sector, "PENDENTE", str(data.get("priority", "NORMAL")).upper(), as_int(data.get("responsible_id")) or None, entry_id, entry["quantity"], user_id, now()),
        )
        add_movement(con, "TAREFAS", cur.lastrowid, "CRIADA_SGO", f"Tarefa criada para {sector} a partir do SGO.", user_id)
        con.commit(); con.close(); return {"ok": True, "task_id": cur.lastrowid}

    @app.get("/api/unified/tasks")
    def list_tasks(sector: str = "", status: str = ""):
        con = db_connect(); sql = """SELECT t.*,u.name responsible_name,c.purchase_id FROM operational_tasks t
                  LEFT JOIN users u ON u.id=t.responsible_id LEFT JOIN cards c ON c.id=t.card_id WHERE 1=1"""; args: list[Any] = []
        if sector: sql += " AND t.sector=?"; args.append(sector)
        if status: sql += " AND t.status=?"; args.append(status)
        sql += " ORDER BY CASE t.priority WHEN 'URGENTE' THEN 0 WHEN 'ALTA' THEN 1 ELSE 2 END,t.id DESC"
        rows = [dict(r) for r in con.execute(sql,args).fetchall()]; con.close(); return rows

    @app.post("/api/unified/tasks")
    async def create_task(request: Request):
        data = await request.json(); user_id = as_int(data.get("user_id"))
        con = db_connect(); require_user(con,user_id)
        title = str(data.get("title", "")).strip(); sector = str(data.get("sector", "")).upper().strip()
        if not title or not sector: con.close(); raise HTTPException(400, "Título e setor são obrigatórios.")
        cur = con.execute(
            """INSERT INTO operational_tasks(title,description,sector,status,priority,responsible_id,card_id,expected_qty,created_by,created_at)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (title,str(data.get("description","")),sector,"PENDENTE",str(data.get("priority","NORMAL")).upper(),as_int(data.get("responsible_id")) or None,as_int(data.get("card_id")) or None,as_int(data.get("expected_qty")),user_id,now()),
        )
        add_movement(con,"TAREFAS",cur.lastrowid,"CRIADA",f"Tarefa '{title}' criada para {sector}.",user_id)
        con.commit();con.close();return {"ok":True,"task_id":cur.lastrowid}

    @app.patch("/api/unified/tasks/{task_id}")
    async def update_task(task_id: int, request: Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id)
        task=con.execute("SELECT * FROM operational_tasks WHERE id=?",(task_id,)).fetchone()
        if not task: con.close();raise HTTPException(404,"Tarefa não encontrada.")
        status=str(data.get("status",task["status"])).upper();responsible=as_int(data.get("responsible_id")) or task["responsible_id"]
        completed=now() if status=="CONCLUIDA" else None
        con.execute("UPDATE operational_tasks SET status=?,responsible_id=?,completed_at=? WHERE id=?",(status,responsible,completed,task_id))
        add_movement(con,"TAREFAS",task_id,"STATUS",f"Tarefa atualizada para {status}.",user_id)
        con.commit();con.close();return {"ok":True}

    @app.get("/api/unified/shipments")
    def list_shipments():
        con=db_connect();rows=[dict(r) for r in con.execute(
            """SELECT s.*,
                      COALESCE((SELECT SUM(si.checked_qty) FROM shipment_items si WHERE si.shipment_id=s.id),0) checked_qty,
                      COALESCE((SELECT COUNT(*) FROM shipment_items si WHERE si.shipment_id=s.id),0) item_count
               FROM shipments s ORDER BY s.id DESC"""
        ).fetchall()];con.close();return rows

    @app.get("/api/unified/shipments/{shipment_id}")
    def shipment_detail(shipment_id:int):
        con=db_connect();shipment=con.execute(
            """SELECT s.*,
                      COALESCE((SELECT SUM(si.checked_qty) FROM shipment_items si WHERE si.shipment_id=s.id),0) checked_qty,
                      COALESCE((SELECT COUNT(*) FROM shipment_items si WHERE si.shipment_id=s.id),0) item_count
               FROM shipments s WHERE s.id=?""",(shipment_id,)
        ).fetchone()
        if not shipment: con.close();raise HTTPException(404,"Romaneio não encontrado.")
        items=[dict(r) for r in con.execute("SELECT * FROM shipment_items WHERE shipment_id=? ORDER BY id",(shipment_id,)).fetchall()]
        result={"shipment":dict(shipment),"items":items};con.close();return result

    @app.post("/api/unified/shipments")
    async def create_shipment(request: Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","estocagem","expedicao"})
        doc=str(data.get("document_no","")).strip();items=data.get("items") or []
        if not doc: con.close();raise HTTPException(400,"Informe o número do romaneio.")
        total=sum(max(0,as_int(i.get("quantity"))) for i in items)
        status="EM_CONFERENCIA" if total>0 else "RASCUNHO"
        cur=con.execute(
            """INSERT INTO shipments(document_no,destination,status,total_qty,created_by,created_at,
               carrier,vehicle_plate,driver_name,scheduled_at,volume_count,notes,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (doc,str(data.get("destination","")).strip(),status,total,user_id,now(),str(data.get("carrier","")).strip(),
             str(data.get("vehicle_plate","")).strip().upper(),str(data.get("driver_name","")).strip(),
             str(data.get("scheduled_at","")).strip(),max(0,as_int(data.get("volume_count"))),str(data.get("notes","")).strip(),now())
        )
        for item in items:
            qty=max(0,as_int(item.get("quantity")))
            if qty<=0: continue
            con.execute("INSERT INTO shipment_items(shipment_id,stock_entry_id,barcode,description,quantity,checked_qty,notes) VALUES(?,?,?,?,?,?,?)",(cur.lastrowid,as_int(item.get("stock_entry_id")) or None,str(item.get("barcode","")).strip(),str(item.get("description","")).strip(),qty,min(qty,max(0,as_int(item.get("checked_qty")))),str(item.get("notes","")).strip()))
        add_movement(con,"EXPEDICAO",cur.lastrowid,"ROMANEIO_CRIADO",f"Romaneio {doc} criado com {total} peça(s).",user_id)
        con.commit();con.close();return {"ok":True,"shipment_id":cur.lastrowid}

    @app.patch("/api/unified/shipments/{shipment_id}/filling")
    async def update_shipment_filling(shipment_id:int,request:Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","estocagem","expedicao"})
        shipment=con.execute("SELECT * FROM shipments WHERE id=?",(shipment_id,)).fetchone()
        if not shipment: con.close();raise HTTPException(404,"Romaneio não encontrado.")
        if shipment["status"]=="EXPEDIDO": con.close();raise HTTPException(400,"Um romaneio expedido não pode mais ser alterado.")
        fields={
            "document_no":str(data.get("document_no",shipment["document_no"] or "")).strip(),
            "destination":str(data.get("destination",shipment["destination"] or "")).strip(),
            "carrier":str(data.get("carrier",shipment["carrier"] or "")).strip(),
            "vehicle_plate":str(data.get("vehicle_plate",shipment["vehicle_plate"] or "")).strip().upper(),
            "driver_name":str(data.get("driver_name",shipment["driver_name"] or "")).strip(),
            "scheduled_at":str(data.get("scheduled_at",shipment["scheduled_at"] or "")).strip(),
            "volume_count":max(0,as_int(data.get("volume_count",shipment["volume_count"]))),
            "notes":str(data.get("notes",shipment["notes"] or "")).strip(),
        }
        if not fields["document_no"]: con.close();raise HTTPException(400,"Informe o número do romaneio.")
        for item in data.get("items") or []:
            item_id=as_int(item.get("id"));row=con.execute("SELECT quantity FROM shipment_items WHERE id=? AND shipment_id=?",(item_id,shipment_id)).fetchone()
            if not row: continue
            checked=max(0,as_int(item.get("checked_qty")))
            if checked>row["quantity"]: con.close();raise HTTPException(400,"A quantidade conferida não pode superar a prevista.")
            con.execute("UPDATE shipment_items SET checked_qty=?,notes=? WHERE id=?",(checked,str(item.get("notes","")).strip(),item_id))
        total=con.execute("SELECT COALESCE(SUM(quantity),0) n FROM shipment_items WHERE shipment_id=?",(shipment_id,)).fetchone()["n"]
        checked=con.execute("SELECT COALESCE(SUM(checked_qty),0) n FROM shipment_items WHERE shipment_id=?",(shipment_id,)).fetchone()["n"]
        required_complete=bool(fields["destination"] and fields["carrier"] and fields["vehicle_plate"] and fields["driver_name"] and fields["scheduled_at"] and fields["volume_count"]>0)
        status="PRONTO" if required_complete and total>0 and checked==total else ("EM_CONFERENCIA" if total>0 else "RASCUNHO")
        con.execute(
            """UPDATE shipments SET document_no=?,destination=?,carrier=?,vehicle_plate=?,driver_name=?,scheduled_at=?,
               volume_count=?,notes=?,total_qty=?,status=?,updated_at=? WHERE id=?""",
            (fields["document_no"],fields["destination"],fields["carrier"],fields["vehicle_plate"],fields["driver_name"],fields["scheduled_at"],fields["volume_count"],fields["notes"],total,status,now(),shipment_id)
        )
        add_movement(con,"EXPEDICAO",shipment_id,"PREENCHIMENTO",f"Preenchimento salvo: {checked}/{total} peça(s), status {status}.",user_id)
        con.commit();con.close();return {"ok":True,"status":status,"checked_qty":checked,"total_qty":total}

    @app.patch("/api/unified/shipments/{shipment_id}")
    async def update_shipment(shipment_id:int,request:Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","estocagem","expedicao"})
        status=str(data.get("status","EXPEDIDO")).upper();shipment=con.execute("SELECT * FROM shipments WHERE id=?",(shipment_id,)).fetchone()
        if not shipment: con.close();raise HTTPException(404,"Romaneio não encontrado.")
        if status=="EXPEDIDO":
            checked=con.execute("SELECT COALESCE(SUM(checked_qty),0) n FROM shipment_items WHERE shipment_id=?",(shipment_id,)).fetchone()["n"]
            missing=[label for label,value in [("destino",shipment["destination"]),("transportadora",shipment["carrier"]),("placa",shipment["vehicle_plate"]),("motorista",shipment["driver_name"]),("previsão de saída",shipment["scheduled_at"]),("volumes",shipment["volume_count"])] if not value]
            if missing or shipment["total_qty"]<=0 or checked!=shipment["total_qty"]:
                con.close();raise HTTPException(400,"Preenchimento incompleto. Revise dados do transporte e conferência de todas as peças.")
        con.execute("UPDATE shipments SET status=?,completed_at=?,updated_at=? WHERE id=?",(status,now() if status=="EXPEDIDO" else None,now(),shipment_id))
        add_movement(con,"EXPEDICAO",shipment_id,"STATUS",f"Expedição atualizada para {status}.",user_id);con.commit();con.close();return {"ok":True}

    @app.post("/api/unified/shipments/import-pdf")
    async def import_shipment_pdf(user_id: int, file: UploadFile = File(...)):
        raw = await file.read(); parsed_doc, items = parse_pdf_items(raw)
        con=db_connect();require_user(con,user_id,{"admin","supervisor","estocagem","expedicao"})
        doc=parsed_doc or Path(file.filename or "romaneio").stem
        total=sum(as_int(i["quantity"]) for i in items)
        cur=con.execute("INSERT INTO shipments(document_no,destination,status,total_qty,created_by,created_at) VALUES(?,?,?,?,?,?)",(doc,"","PREPARANDO",total,user_id,now()))
        for item in items:
            con.execute("INSERT INTO shipment_items(shipment_id,barcode,description,quantity) VALUES(?,?,?,?)",(cur.lastrowid,item["barcode"],item["description"],item["quantity"]))
        add_movement(con,"EXPEDICAO",cur.lastrowid,"PDF_IMPORTADO",f"Romaneio {doc} importado do PDF com {len(items)} item(ns) e {total} peça(s).",user_id)
        con.commit();con.close();return {"ok":True,"shipment_id":cur.lastrowid,"document_no":doc,"items":len(items),"total_qty":total}

    @app.get("/api/unified/returns")
    def list_returns(status:str="",search:str=""):
        con=db_connect();sql="SELECT * FROM returns WHERE 1=1";args=[]
        if status:sql+=" AND status=?";args.append(status)
        if search:q=f"%{search}%";sql+=" AND (document_no LIKE ? OR store LIKE ? OR customer LIKE ?)";args += [q,q,q]
        sql+=" ORDER BY id DESC";rows=[dict(r) for r in con.execute(sql,args).fetchall()];con.close();return rows

    @app.get("/api/unified/returns/{return_id}")
    def return_detail(return_id:int):
        con=db_connect();head=con.execute("SELECT * FROM returns WHERE id=?",(return_id,)).fetchone()
        if not head:con.close();raise HTTPException(404,"Devolução não encontrada.")
        items=[dict(r) for r in con.execute("SELECT * FROM return_items WHERE return_id=? ORDER BY id",(return_id,)).fetchall()];con.close();return {"return":dict(head),"items":items}

    @app.post("/api/unified/returns")
    async def create_return(request:Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","devolucoes"})
        doc=str(data.get("document_no","")).strip();items=data.get("items") or []
        if not doc:con.close();raise HTTPException(400,"Informe o documento da devolução.")
        store_total=sum(as_int(i.get("qty_store")) for i in items);cd_total=sum(as_int(i.get("qty_cd")) for i in items);ana_total=sum(as_int(i.get("qty_anapolis")) for i in items)
        difference=cd_total+ana_total-store_total;status="DIVERGENTE" if difference or any((as_int(i.get("qty_cd"))+as_int(i.get("qty_anapolis"))-as_int(i.get("qty_store"))) for i in items) else "AGUARDANDO_TRATAMENTO"
        cur=con.execute("""INSERT INTO returns(document_no,store,customer,status,total_store,total_cd,total_anapolis,difference,notes,created_by,created_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)""",(doc,str(data.get("store","")),str(data.get("customer","")),status,store_total,cd_total,ana_total,difference,str(data.get("notes","")),user_id,now()))
        rid=cur.lastrowid
        for item in items:
            qs=as_int(item.get("qty_store"));qc=as_int(item.get("qty_cd"));qa=as_int(item.get("qty_anapolis"));diff=qc+qa-qs
            con.execute("""INSERT INTO return_items(return_id,barcode,reference,description,size,qty_store,qty_cd,qty_anapolis,difference,condition_status,destination,notes)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",(rid,str(item.get("barcode","")),str(item.get("reference","")),str(item.get("description","")),str(item.get("size","")),qs,qc,qa,diff,"OK" if diff==0 else "DIVERGENTE","AVARIA" if qa else None,str(item.get("notes",""))))
        add_movement(con,"DEVOLUCOES",rid,"REGISTRADA",f"Devolução {doc} registrada com {store_total} peça(s).",user_id)
        con.commit();con.close();return {"ok":True,"return_id":rid,"status":status}

    @app.patch("/api/unified/returns/{return_id}")
    async def update_return(return_id:int,request:Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","devolucoes"})
        status=str(data.get("status","")).upper();decision=str(data.get("decision","")).upper()
        if not status:con.close();raise HTTPException(400,"Informe o novo status.")
        con.execute("UPDATE returns SET status=?,decision=?,completed_at=? WHERE id=?",(status,decision,now() if status=="CONCLUIDA" else None,return_id))
        add_movement(con,"DEVOLUCOES",return_id,"STATUS",f"Devolução atualizada para {status}.",user_id);con.commit();con.close();return {"ok":True}

    @app.post("/api/unified/returns/compare-pdf")
    async def compare_return_pdfs(
        user_id: int,
        file_store: UploadFile = File(...),
        file_cd: UploadFile = File(...),
        file_anapolis: UploadFile | None = File(None),
    ):
        store_doc, store_items = parse_pdf_items(await file_store.read())
        cd_doc, cd_items = parse_pdf_items(await file_cd.read())
        ana_doc, ana_items = ("", [])
        if file_anapolis and file_anapolis.filename:
            ana_doc, ana_items = parse_pdf_items(await file_anapolis.read())
        by_source = [{i["barcode"]: i for i in source} for source in (store_items, cd_items, ana_items)]
        barcodes = sorted(set().union(*(set(source) for source in by_source)))
        items = []
        for barcode in barcodes:
            candidates = [source.get(barcode) for source in by_source]
            description = next((item["description"] for item in candidates if item), "Item importado do PDF")
            items.append({"barcode":barcode,"description":description,"qty_store":candidates[0]["quantity"] if candidates[0] else 0,"qty_cd":candidates[1]["quantity"] if candidates[1] else 0,"qty_anapolis":candidates[2]["quantity"] if candidates[2] else 0})
        con=db_connect();require_user(con,user_id,{"admin","supervisor","devolucoes"})
        store_total=sum(i["qty_store"] for i in items);cd_total=sum(i["qty_cd"] for i in items);ana_total=sum(i["qty_anapolis"] for i in items)
        difference=cd_total+ana_total-store_total
        status="DIVERGENTE" if difference or any(i["qty_cd"]+i["qty_anapolis"]-i["qty_store"] for i in items) else "AGUARDANDO_TRATAMENTO"
        doc=store_doc or cd_doc or ana_doc or Path(file_store.filename or "devolucao").stem
        cur=con.execute("""INSERT INTO returns(document_no,store,customer,status,total_store,total_cd,total_anapolis,difference,notes,created_by,created_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)""",(doc,"","",status,store_total,cd_total,ana_total,difference,"Comparação automática de PDFs; revisar antes da conclusão.",user_id,now()))
        rid=cur.lastrowid
        for item in items:
            diff=item["qty_cd"]+item["qty_anapolis"]-item["qty_store"]
            con.execute("""INSERT INTO return_items(return_id,barcode,description,qty_store,qty_cd,qty_anapolis,difference,condition_status,destination,notes)
                         VALUES(?,?,?,?,?,?,?,?,?,?)""",(rid,item["barcode"],item["description"],item["qty_store"],item["qty_cd"],item["qty_anapolis"],diff,"OK" if diff==0 else "DIVERGENTE","AVARIA" if item["qty_anapolis"] else None,""))
        add_movement(con,"DEVOLUCOES",rid,"PDFS_COMPARADOS",f"Devolução {doc}: Loja {store_total}, CD {cd_total}, Anápolis {ana_total}.",user_id)
        con.commit();con.close();return {"ok":True,"return_id":rid,"document_no":doc,"items":len(items),"status":status,"difference":difference}

    @app.patch("/api/unified/returns/{return_id}/items/{item_id}")
    async def treat_return_item(return_id:int,item_id:int,request:Request):
        data=await request.json();user_id=as_int(data.get("user_id"));con=db_connect();require_user(con,user_id,{"admin","supervisor","devolucoes"})
        item=con.execute("SELECT * FROM return_items WHERE id=? AND return_id=?",(item_id,return_id)).fetchone()
        if not item:con.close();raise HTTPException(404,"Item da devolução não encontrado.")
        condition=str(data.get("condition_status",item["condition_status"])).upper();destination=str(data.get("destination",item["destination"] or "")).upper();notes=str(data.get("notes",item["notes"] or ""))
        con.execute("UPDATE return_items SET condition_status=?,destination=?,notes=? WHERE id=?",(condition,destination,notes,item_id))
        pending=con.execute("SELECT COUNT(*) n FROM return_items WHERE return_id=? AND condition_status IN ('PENDENTE','DIVERGENTE')",(return_id,)).fetchone()["n"]
        if not pending:con.execute("UPDATE returns SET status='AGUARDANDO_CONCLUSAO' WHERE id=? AND status!='CONCLUIDA'",(return_id,))
        add_movement(con,"DEVOLUCOES",return_id,"ITEM_TRATADO",f"Item {item['barcode'] or item_id} destinado para {destination or 'revisão'}.",user_id)
        con.commit();con.close();return {"ok":True,"pending_items":pending}

    @app.get("/api/unified/movements")
    def movements(limit:int=200):
        con=db_connect();rows=[dict(r) for r in con.execute("""SELECT m.*,u.name user_name FROM unified_movements m LEFT JOIN users u ON u.id=m.user_id ORDER BY m.id DESC LIMIT ?""",(max(1,min(limit,1000)),)).fetchall()];con.close();return rows

    @app.get("/api/unified/warehouse/heatmap")
    def warehouse_heatmap():
        """Ocupação por Rua/coluna, sem alterar os lançamentos de estoque."""
        con = db_connect()
        rows = con.execute(
            """SELECT l.aisle,l.column_no,l.side,SUM(l.capacity) capacity,
                      SUM(l.occupied_qty) occupied,COUNT(*) niveis
               FROM warehouse_locations l JOIN warehouse_zones z ON z.id=l.zone_id
               WHERE z.active=1
               GROUP BY l.aisle,l.column_no,l.side ORDER BY l.aisle,l.column_no"""
        ).fetchall()
        con.close()
        sequential_columns = {
            rua: set(config.get("cols_seq", [])) for rua, config in ESTRUTURA_CD.items()
        }
        ruas: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            capacity = int(row["capacity"] or 0)
            occupied = int(row["occupied"] or 0)
            section = (
                "sequencial"
                if row["column_no"] in sequential_columns.get(row["aisle"], set())
                else row["side"]
            )
            ruas.setdefault(row["aisle"], []).append({
                "coluna": row["column_no"], "lado": row["side"], "secao": section,
                "niveis": row["niveis"], "capacidade": capacity, "ocupado": occupied,
                "ocupacao_pct": round(occupied * 100 / capacity, 1) if capacity else 0.0,
            })
        return {
            "ruas": [
                {"rua": rua, "colunas": columns}
                for rua, columns in sorted(ruas.items(), key=lambda item: int(item[0].split()[1]))
            ]
        }
