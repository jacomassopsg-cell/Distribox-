from __future__ import annotations

import json
import math
import os
import re
import sqlite3
import unicodedata
import uuid
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from quality_module import init_quality_db, quality_card_data, register_quality_routes
from processing_module import init_processing_db, processing_card_data, register_processing_routes
from downstream_module import downstream_card_data, init_downstream_db, register_downstream_routes
from unified_module import init_unified_db, register_unified_routes

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("OUTLOG_DATA_DIR", str(BASE_DIR / "data"))).resolve()
UPLOAD_DIR = DATA_DIR / "uploads"
DB_PATH = DATA_DIR / "controle_logistica.db"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="OutLog One - Operação Logística Integrada", version="3.1.0")


@app.middleware("http")
async def disable_stale_interface_cache(request: Request, call_next):
    response = await call_next(request)
    if request.url.path == "/" or request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

KANBAN_TRANSITO = "1.3 Compras - Em Trânsito"

STATUS_LABELS = {
    "AGUARDANDO_RECEBIMENTO": "Aguardando recebimento",
    "RECEBIMENTO_FISICO_CONCLUIDO_10_PENDENTE": "Recebimento físico concluído — 10% pendente",
    "SEPARACAO_10_CONCLUIDA_RECEBIMENTO_PENDENTE": "10% concluído — recebimento físico pendente",
    "RECEBIMENTO_CONCLUIDO": "Recebimento concluído",
    "AGUARDANDO_QUALIDADE": "Aguardando inspeção",
    "AGUARDANDO_DESPACHO_COSTURA": "Aguardando despacho para Costura",
    "EM_COSTURA_CD01": "Em Costura",
    "DESPACHO_CD02": "Despacho CD02",
    "AGUARDANDO_RECEBIMENTO_RETORNO": "Aguardando recebimento do retorno",
    "RETORNO_CONCLUIDO": "Retorno concluído — aguardando inspeção",
    "RETORNO_FISICO_CONCLUIDO_10_PENDENTE": "Retorno recebido — separação dos 10% pendente",
    "SEPARACAO_RETORNO_10_CONCLUIDA_FISICO_PENDENTE": "10% do retorno separados — recebimento pendente",
    "AGUARDANDO_INSPECAO": "Aguardando inspeção",
    "EM_INSPECAO": "Em inspeção",
    "INSPECAO_PAUSADA": "Inspeção pausada",
    "AGUARDANDO_CONCLUSAO_QUALIDADE": "Aguardando conclusão da Qualidade",
    "AGUARDANDO_PROCESSAMENTO": "Aguardando Processamento",
    "AGUARDANDO_INICIO_PROCESSAMENTO": "Aguardando início do Processamento",
    "EM_PROCESSAMENTO": "Em Processamento",
    "PROCESSAMENTO_PAUSADO": "Processamento pausado",
    "AGUARDANDO_CONCLUSAO_PROCESSAMENTO": "Aguardando conclusão do Processamento",
    "AGUARDANDO_TRIAGEM": "Aguardando Triagem",
    "AGUARDANDO_ETIQUETAGEM": "Aguardando Etiquetagem",
    "AGUARDANDO_ESTOCAGEM": "Aguardando Estocagem",
    "FINALIZADO": "Finalizado",
    "ORIGEM_FORNECEDOR": "No fornecedor",
    "ORIGEM_TRANSITO": "Em trânsito",
    "ORIGEM_QUALIDADE": "Na Qualidade",
    "ORIGEM_PCP_CONFIGURAR": "PCP — configurar aviamento",
    "ORIGEM_AGUARDANDO_COSTURA": "Aguardando envio à Costura",
    "ORIGEM_COSTURA": "Em Costura",
    "ORIGEM_RETORNO_COSTURA": "Retornou da Costura",
    "ORIGEM_PROCESSAMENTO": "Área de Processamento",
    "ORIGEM_ESTOCAGEM": "Em Estocagem",
    "ORIGEM_MISTO": "Itens em etapas diferentes",
}

RECEIVING_STATUSES = {
    "AGUARDANDO_RECEBIMENTO",
    "RECEBIMENTO_FISICO_CONCLUIDO_10_PENDENTE",
    "SEPARACAO_10_CONCLUIDA_RECEBIMENTO_PENDENTE",
    "AGUARDANDO_DESPACHO_COSTURA",
    "AGUARDANDO_RECEBIMENTO_RETORNO",
    "RETORNO_FISICO_CONCLUIDO_10_PENDENTE",
    "SEPARACAO_RETORNO_10_CONCLUIDA_FISICO_PENDENTE",
}


def db_connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys=ON")
    return con


def iso_now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text.lower())
        if unicodedata.category(ch) != "Mn"
    )


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_text(value))


def kanban_matches(value: Any) -> bool:
    return normalize_text(value) == normalize_text(KANBAN_TRANSITO)


def purchase_mode_from_type(value: Any) -> Optional[str]:
    text = normalize_text(value)
    if "private label" in text or "privatelabel" in text:
        return "GRADE"
    if "saldo" in text:
        return "SALDO"
    return None


def source_stage_from_row(values: dict[str, Any]) -> str:
    kanban = normalize_text(values.get("status_kanban"))
    logistics = normalize_text(values.get("status_logistics"))
    pcp = normalize_text(values.get("status_pcp"))
    quality = normalize_text(values.get("status_quality"))
    lot = normalize_text(values.get("status_lot"))
    purchase = normalize_text(values.get("status_purchase"))
    phase = normalize_text(values.get("inspection_phase"))
    if "em estocagem" in logistics or kanban.startswith("5.3"):
        return "ESTOCAGEM"
    if "area de processamento" in logistics or kanban.startswith("5.2"):
        return "PROCESSAMENTO"
    if "aguardando processamento" in logistics or kanban.startswith("5.1"):
        return "AGUARDANDO_PROCESSAMENTO"
    if "em aviamento" in pcp or kanban.startswith("3.3"):
        return "EM_COSTURA"
    if "aguardando envio" in pcp or kanban.startswith("3.2"):
        return "AGUARDANDO_COSTURA"
    if "configurar aviamento" in pcp or kanban.startswith("3.1"):
        return "PCP_CONFIGURAR"
    if "concluido" in pcp and phase == "customization":
        return "RETORNO_COSTURA"
    if "retrabalho" in quality or "retrabalho" in lot or kanban.startswith("2.0"):
        return "QUALIDADE_RETRABALHO"
    if "reprovado" in quality or "rejeitado" in lot or kanban.startswith("9.2"):
        return "QUALIDADE_REJEITADO"
    if "inspecao" in quality or "aguardando inspecao" in lot or kanban.startswith("2.1"):
        return "QUALIDADE"
    if kanban.startswith("1.3") or "transito" in purchase:
        return "TRANSITO"
    if kanban.startswith("1.2") or "fornecedor" in purchase:
        return "FORNECEDOR"
    if kanban.startswith("6.0") or "concluido" in logistics or "concluido" in purchase:
        return "CONCLUIDO"
    return "NAO_CLASSIFICADO"


def card_source_status(stages: set[str]) -> tuple[str, str]:
    active = {stage for stage in stages if stage not in {"CONCLUIDO", "NAO_CLASSIFICADO"}}
    if len(active) > 1:
        return "ORIGEM_MISTO", " / ".join(sorted(active))
    stage = next(iter(active or stages or {"NAO_CLASSIFICADO"}))
    mapping = {
        "FORNECEDOR":"ORIGEM_FORNECEDOR", "TRANSITO":"ORIGEM_TRANSITO",
        "QUALIDADE":"ORIGEM_QUALIDADE", "QUALIDADE_RETRABALHO":"ORIGEM_QUALIDADE",
        "QUALIDADE_REJEITADO":"ORIGEM_QUALIDADE", "PCP_CONFIGURAR":"ORIGEM_PCP_CONFIGURAR",
        "AGUARDANDO_COSTURA":"ORIGEM_AGUARDANDO_COSTURA", "EM_COSTURA":"ORIGEM_COSTURA",
        "RETORNO_COSTURA":"ORIGEM_RETORNO_COSTURA", "AGUARDANDO_PROCESSAMENTO":"AGUARDANDO_PROCESSAMENTO",
        "PROCESSAMENTO":"ORIGEM_PROCESSAMENTO", "ESTOCAGEM":"ORIGEM_ESTOCAGEM",
        "CONCLUIDO":"FINALIZADO", "NAO_CLASSIFICADO":"AGUARDANDO_RECEBIMENTO",
    }
    return mapping.get(stage, "ORIGEM_MISTO"), stage


def imported_card_route(stages: set[str]) -> tuple[str, str]:
    """O Recebimento mantém a visão-mãe; os demais setores usam as etapas dos itens."""
    status, _ = card_source_status(stages)
    return "RECEBIMENTO", status


def is_cd02(value: Any) -> bool:
    return re.sub(r"[^a-z0-9]", "", normalize_text(value)) in {"cd2", "cd02"}


def delete_card_with_downstream(con: sqlite3.Connection, card_id: int) -> None:
    """Remove primeiro os registros V8, cujas tabelas antigas não possuíam cascade."""
    operation_ids = [row["id"] for row in con.execute(
        "SELECT id FROM downstream_operations WHERE card_id=?", (card_id,)
    ).fetchall()]
    if operation_ids:
        marks = ",".join("?" for _ in operation_ids)
        con.execute(f"DELETE FROM downstream_assignments WHERE operation_id IN ({marks})", operation_ids)
        con.execute(f"DELETE FROM downstream_operations WHERE id IN ({marks})", operation_ids)
    con.execute("DELETE FROM cards WHERE id=?", (card_id,))


def has_local_workflow(con: sqlite3.Connection, card_id: int, current_sector: str) -> bool:
    if current_sector != "RECEBIMENTO":
        return True
    checks = [
        "SELECT 1 FROM quality_inspections WHERE card_id=? LIMIT 1",
        "SELECT 1 FROM processing_records WHERE card_id=? LIMIT 1",
        "SELECT 1 FROM downstream_operations WHERE card_id=? LIMIT 1",
        "SELECT 1 FROM dispatches WHERE card_id=? LIMIT 1",
        """SELECT 1 FROM receivings r WHERE r.card_id=? AND
             (r.physical_status!='PENDENTE' OR r.ten_percent_status!='PENDENTE') LIMIT 1""",
    ]
    return any(con.execute(sql, (card_id,)).fetchone() for sql in checks)


def clean_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(round(float(value)))
    except Exception:
        return 0


def clean_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() or None


def to_json(value: Any) -> str:
    return json.dumps(value or [], ensure_ascii=False)


def from_json(value: Optional[str], default: Any) -> Any:
    if not value:
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def business_seconds_between(start: datetime, end: datetime) -> int:
    if end <= start:
        return 0
    total = 0
    day = start.date()
    while day <= end.date():
        weekday = day.weekday()
        if weekday <= 3:
            work_start = datetime.combine(day, time(8, 0))
            work_end = datetime.combine(day, time(18, 0))
        elif weekday == 4:
            work_start = datetime.combine(day, time(8, 0))
            work_end = datetime.combine(day, time(17, 0))
        else:
            day += timedelta(days=1)
            continue
        interval_start = max(start, work_start)
        interval_end = min(end, work_end)
        if interval_end > interval_start:
            total += int((interval_end - interval_start).total_seconds())
        day += timedelta(days=1)
    return total


def add_business_days(start_date: date, days: int) -> date:
    current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


def init_db() -> None:
    con = db_connect()
    con.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS cards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id TEXT UNIQUE NOT NULL,
            source_created_date TEXT,
            supplier TEXT,
            original_type TEXT,
            purchase_mode TEXT,
            status_compra TEXT,
            original_destination TEXT,
            forecast_date TEXT,
            qtd_itens INTEGER DEFAULT 0,
            source_notes TEXT,
            brand TEXT,
            collection TEXT,
            current_sector TEXT NOT NULL DEFAULT 'RECEBIMENTO',
            status TEXT NOT NULL DEFAULT 'AGUARDANDO_RECEBIMENTO',
            receiving_type TEXT NOT NULL DEFAULT 'NOVA',
            quality_destination TEXT,
            casulo_current TEXT,
            source_location_summary TEXT,
            source_snapshot_at TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            source_key TEXT NOT NULL,
            product TEXT,
            reference TEXT,
            sku TEXT,
            group_name TEXT,
            collection TEXT,
            brand TEXT,
            gender TEXT,
            color TEXT,
            size TEXT,
            capsule TEXT,
            lot TEXT,
            lot_id TEXT,
            nf TEXT,
            expected_qty INTEGER NOT NULL DEFAULT 0,
            url_photo TEXT,
            status_kanban TEXT,
            source_stage TEXT,
            source_status_purchase TEXT,
            source_status_lot TEXT,
            source_status_quality TEXT,
            source_inspection_phase TEXT,
            source_status_pcp TEXT,
            source_seamstress TEXT,
            source_status_logistics TEXT,
            source_received_qty INTEGER DEFAULT 0,
            UNIQUE(card_id, source_key),
            FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            total_rows INTEGER NOT NULL,
            matched_rows INTEGER NOT NULL,
            cards_created INTEGER NOT NULL,
            cards_updated INTEGER NOT NULL,
            items_created INTEGER NOT NULL,
            items_updated INTEGER NOT NULL,
            errors TEXT,
            user_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS receivings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            receiving_type TEXT NOT NULL,
            physical_status TEXT NOT NULL DEFAULT 'PENDENTE',
            volumes INTEGER,
            received_qty INTEGER,
            has_damage INTEGER,
            damage_description TEXT,
            notes TEXT,
            photo_paths TEXT,
            ten_percent_required INTEGER NOT NULL DEFAULT 0,
            ten_percent_min INTEGER NOT NULL DEFAULT 0,
            ten_percent_actual INTEGER,
            ten_percent_status TEXT NOT NULL DEFAULT 'NAO_INICIADA',
            triage_completed INTEGER NOT NULL DEFAULT 0,
            physical_completed_by INTEGER,
            physical_completed_at TEXT,
            closed_at TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE,
            FOREIGN KEY(physical_completed_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS timer_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receiving_id INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            event_at TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            FOREIGN KEY(receiving_id) REFERENCES receivings(id) ON DELETE CASCADE,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS dispatches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            destination TEXT NOT NULL,
            carrier TEXT,
            seamstress_name TEXT,
            dispatched_qty INTEGER,
            volumes INTEGER,
            photo_paths TEXT,
            notes TEXT,
            return_forecast TEXT,
            completed_by INTEGER,
            completed_at TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE,
            FOREIGN KEY(completed_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS casulo_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            old_casulo TEXT,
            new_casulo TEXT NOT NULL,
            note TEXT,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            user_id INTEGER,
            event_type TEXT NOT NULL,
            description TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(card_id) REFERENCES cards(id) ON DELETE CASCADE,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        """
    )
    for username, password, name, role in [
        ("admin", "1234", "Administrador", "admin"),
        ("recebimento", "1234", "Operador do Recebimento", "recebimento"),
        ("recebimento2", "1234", "Segundo Operador do Recebimento", "recebimento"),
        ("consulta", "1234", "Usuário de Consulta", "consulta"),
        ("qualidade1", "1234", "Inspetor da Qualidade 1", "qualidade"),
        ("qualidade2", "1234", "Inspetor da Qualidade 2", "qualidade"),
        ("supervisor", "1234", "Supervisor", "supervisor"),
        ("processamento1", "1234", "Operador do Processamento 1", "processamento"),
        ("processamento2", "1234", "Operador do Processamento 2", "processamento"),
        ("etiquetagem1", "1234", "Operador da Etiquetagem 1", "etiquetagem"),
        ("etiquetagem2", "1234", "Operador da Etiquetagem 2", "etiquetagem"),
        ("estocagem1", "1234", "Operador da Estocagem 1", "estocagem"),
        ("estocagem2", "1234", "Operador da Estocagem 2", "estocagem"),
        ("planejamento1", "1234", "Planejamento SGO", "planejamento"),
        ("expedicao1", "1234", "Operador da Expedição", "expedicao"),
        ("devolucoes1", "1234", "Operador de Devoluções", "devolucoes"),
    ]:
        con.execute(
            "INSERT OR IGNORE INTO users(username,password,name,role) VALUES(?,?,?,?)",
            (username, password, name, role),
        )
    # Migração funcional V4: Cards CD01 permanecem na aba Recebimento enquanto estão em Costura.
    con.execute(
        "UPDATE cards SET current_sector='RECEBIMENTO' WHERE status='EM_COSTURA_CD01'"
    )
    card_columns = {row["name"] for row in con.execute("PRAGMA table_info(cards)").fetchall()}
    if "purchase_mode" not in card_columns:
        con.execute("ALTER TABLE cards ADD COLUMN purchase_mode TEXT")
    for name, definition in [("source_location_summary","TEXT"),("source_snapshot_at","TEXT")]:
        if name not in card_columns:
            con.execute(f"ALTER TABLE cards ADD COLUMN {name} {definition}")
    item_columns = {row["name"] for row in con.execute("PRAGMA table_info(items)").fetchall()}
    for name, definition in [
        ("source_stage","TEXT"),("source_status_purchase","TEXT"),("source_status_lot","TEXT"),
        ("source_status_quality","TEXT"),("source_inspection_phase","TEXT"),("source_status_pcp","TEXT"),
        ("source_seamstress","TEXT"),("source_status_logistics","TEXT"),("source_received_qty","INTEGER DEFAULT 0"),
    ]:
        if name not in item_columns:
            con.execute(f"ALTER TABLE items ADD COLUMN {name} {definition}")
    receiving_columns = {row["name"] for row in con.execute("PRAGMA table_info(receivings)").fetchall()}
    if "triage_completed" not in receiving_columns:
        con.execute("ALTER TABLE receivings ADD COLUMN triage_completed INTEGER NOT NULL DEFAULT 0")
    con.execute("UPDATE receivings SET ten_percent_required=1 WHERE receiving_type='RETORNO' AND closed_at IS NULL")
    open_returns = con.execute(
        """SELECT r.id,r.received_qty,c.purchase_mode,c.id card_id
           FROM receivings r JOIN cards c ON c.id=r.card_id
           WHERE r.receiving_type='RETORNO' AND r.closed_at IS NULL"""
    ).fetchall()
    for receiving in open_returns:
        base_qty = int(receiving["received_qty"] or 0)
        item_count = con.execute(
            "SELECT COUNT(*) n FROM items WHERE card_id=? AND expected_qty>0", (receiving["card_id"],)
        ).fetchone()["n"]
        minimum = max(math.ceil(base_qty * 0.10), item_count if receiving["purchase_mode"] == "GRADE" else 0)
        con.execute("UPDATE receivings SET ten_percent_min=? WHERE id=?", (minimum, receiving["id"]))
    con.execute("""UPDATE cards SET purchase_mode=CASE
        WHEN lower(COALESCE(original_type,'')) LIKE '%private%label%' THEN 'GRADE'
        WHEN lower(COALESCE(original_type,'')) LIKE '%saldo%' THEN 'SALDO'
        ELSE purchase_mode END
        WHERE purchase_mode IS NULL OR purchase_mode=''""")
    con.commit()
    con.close()


def add_history(con: sqlite3.Connection, card_id: int, event_type: str, description: str, user_id: Optional[int] = None) -> None:
    con.execute(
        "INSERT INTO history(card_id,user_id,event_type,description,created_at) VALUES(?,?,?,?,?)",
        (card_id, user_id, event_type, description, iso_now()),
    )


def get_user(con: sqlite3.Connection, user_id: int) -> sqlite3.Row:
    user = con.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        raise HTTPException(401, "Usuário inválido.")
    return user


def require_role(con: sqlite3.Connection, user_id: int, allowed: set[str]) -> sqlite3.Row:
    user = get_user(con, user_id)
    if user["role"] not in allowed:
        raise HTTPException(403, "Seu perfil não possui permissão para esta ação.")
    return user


def ensure_receiving(con: sqlite3.Connection, card_id: int, receiving_type: str) -> int:
    row = con.execute(
        "SELECT id FROM receivings WHERE card_id=? AND receiving_type=? ORDER BY id DESC LIMIT 1",
        (card_id, receiving_type),
    ).fetchone()
    if row:
        return row["id"]
    total = con.execute(
        "SELECT COALESCE(SUM(expected_qty),0) total FROM items WHERE card_id=?",
        (card_id,),
    ).fetchone()["total"]
    required = 1
    item_count = con.execute(
        "SELECT COUNT(*) n FROM items WHERE card_id=? AND expected_qty>0", (card_id,)
    ).fetchone()["n"]
    mode = con.execute("SELECT purchase_mode FROM cards WHERE id=?", (card_id,)).fetchone()["purchase_mode"]
    minimum = max(math.ceil(total * 0.10), item_count if mode == "GRADE" else 0)
    cur = con.execute(
        """INSERT INTO receivings(card_id,receiving_type,ten_percent_required,ten_percent_min,created_at)
           VALUES(?,?,?,?,?)""",
        (card_id, receiving_type, required, minimum, iso_now()),
    )
    return cur.lastrowid


def timer_events(con: sqlite3.Connection, receiving_id: int) -> list[sqlite3.Row]:
    return con.execute(
        "SELECT * FROM timer_events WHERE receiving_id=? ORDER BY id",
        (receiving_id,),
    ).fetchall()


def timer_summary(con: sqlite3.Connection, receiving_id: int) -> dict[str, Any]:
    events = timer_events(con, receiving_id)
    state = "NAO_INICIADA"
    first_start: Optional[datetime] = None
    open_start: Optional[datetime] = None
    pause_start: Optional[datetime] = None
    final_end: Optional[datetime] = None
    active_business = 0
    paused_real = 0
    for event in events:
        dt = datetime.fromisoformat(event["event_at"])
        kind = event["event_type"]
        if kind == "START":
            state = "EM_ANDAMENTO"
            first_start = first_start or dt
            open_start = dt
        elif kind == "PAUSE" and open_start:
            active_business += business_seconds_between(open_start, dt)
            open_start = None
            pause_start = dt
            state = "PAUSADA"
        elif kind == "RESUME":
            if pause_start:
                paused_real += int((dt - pause_start).total_seconds())
            pause_start = None
            open_start = dt
            state = "EM_ANDAMENTO"
        elif kind == "FINISH":
            if open_start:
                active_business += business_seconds_between(open_start, dt)
                open_start = None
            if pause_start:
                paused_real += int((dt - pause_start).total_seconds())
                pause_start = None
            final_end = dt
            state = "CONCLUIDA"
    now_dt = datetime.now().replace(microsecond=0)
    if state == "EM_ANDAMENTO" and open_start:
        active_business += business_seconds_between(open_start, now_dt)
    if state == "PAUSADA" and pause_start:
        paused_real += int((now_dt - pause_start).total_seconds())
    permanence = int(((final_end or now_dt) - first_start).total_seconds()) if first_start else 0
    return {
        "state": state,
        "business_seconds": max(active_business, 0),
        "paused_seconds": max(paused_real, 0),
        "permanence_seconds": max(permanence, 0),
        "started_at": first_start.isoformat() if first_start else None,
        "finished_at": final_end.isoformat() if final_end else None,
        "events": [dict(e) for e in events],
    }


def timer_action(con: sqlite3.Connection, receiving_id: int, action: str, user_id: int) -> dict[str, Any]:
    summary = timer_summary(con, receiving_id)
    state = summary["state"]
    map_action = {
        "start": ("NAO_INICIADA", "START"),
        "pause": ("EM_ANDAMENTO", "PAUSE"),
        "resume": ("PAUSADA", "RESUME"),
        "finish": (("EM_ANDAMENTO", "PAUSADA"), "FINISH"),
    }
    if action not in map_action:
        raise HTTPException(400, "Ação inválida.")
    allowed, event_type = map_action[action]
    valid = state in allowed if isinstance(allowed, tuple) else state == allowed
    if not valid:
        raise HTTPException(400, f"Ação incompatível com o estado atual: {state}.")
    con.execute(
        "INSERT INTO timer_events(receiving_id,event_type,event_at,user_id) VALUES(?,?,?,?)",
        (receiving_id, event_type, iso_now(), user_id),
    )
    return timer_summary(con, receiving_id)


def current_receiving(con: sqlite3.Connection, card_id: int) -> Optional[dict[str, Any]]:
    row = con.execute(
        """SELECT r.*,u.name physical_completed_by_name
           FROM receivings r LEFT JOIN users u ON u.id=r.physical_completed_by
           WHERE r.card_id=? ORDER BY r.id DESC LIMIT 1""",
        (card_id,),
    ).fetchone()
    if not row:
        return None
    data = dict(row)
    data["photo_paths"] = from_json(data.get("photo_paths"), [])
    data["timer"] = timer_summary(con, data["id"]) if data["ten_percent_required"] else None
    return data


def current_dispatch(con: sqlite3.Connection, card_id: int) -> Optional[dict[str, Any]]:
    row = con.execute(
        """SELECT d.*,u.name completed_by_name FROM dispatches d
           LEFT JOIN users u ON u.id=d.completed_by
           WHERE d.card_id=? ORDER BY d.id DESC LIMIT 1""",
        (card_id,),
    ).fetchone()
    if not row:
        return None
    data = dict(row)
    data["photo_paths"] = from_json(data.get("photo_paths"), [])
    return data


def update_new_receiving_flow(con: sqlite3.Connection, receiving_id: int, user_id: Optional[int] = None) -> str:
    rec = con.execute("SELECT * FROM receivings WHERE id=?", (receiving_id,)).fetchone()
    if not rec:
        raise HTTPException(404, "Recebimento não encontrado.")
    physical_done = rec["physical_status"] == "CONCLUIDO"
    sample_done = rec["ten_percent_status"] == "CONCLUIDA"
    if physical_done and sample_done:
        is_return = rec["receiving_type"] == "RETORNO"
        status = "RETORNO_CONCLUIDO" if is_return else "AGUARDANDO_QUALIDADE"
        con.execute(
            "UPDATE cards SET current_sector='QUALIDADE',status=?,updated_at=? WHERE id=?",
            (status, iso_now(), rec["card_id"]),
        )
        con.execute("UPDATE receivings SET closed_at=? WHERE id=?", (iso_now(), receiving_id))
        add_history(
            con,
            rec["card_id"],
            "RECEBIMENTO_CONCLUIDO",
            ("Retorno CD01 recebido e nova amostra de 10% separada. Card encaminhado à Inspeção 2."
             if is_return else
             "Recebimento físico, separação dos 10% e triagem inicial concluídos. Card encaminhado automaticamente à Qualidade."),
            user_id,
        )
    elif physical_done:
        status = ("RETORNO_FISICO_CONCLUIDO_10_PENDENTE" if rec["receiving_type"] == "RETORNO"
                  else "RECEBIMENTO_FISICO_CONCLUIDO_10_PENDENTE")
        con.execute(
            "UPDATE cards SET current_sector='RECEBIMENTO',status=?,updated_at=? WHERE id=?",
            (status, iso_now(), rec["card_id"]),
        )
    elif sample_done:
        status = ("SEPARACAO_RETORNO_10_CONCLUIDA_FISICO_PENDENTE" if rec["receiving_type"] == "RETORNO"
                  else "SEPARACAO_10_CONCLUIDA_RECEBIMENTO_PENDENTE")
        con.execute(
            "UPDATE cards SET current_sector='RECEBIMENTO',status=?,updated_at=? WHERE id=?",
            (status, iso_now(), rec["card_id"]),
        )
    else:
        status = "AGUARDANDO_RECEBIMENTO"
    return status


@app.on_event("startup")
def startup() -> None:
    init_db()
    init_quality_db()
    init_processing_db()
    init_downstream_db()
    init_unified_db()


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/api/login")
async def login(request: Request):
    data = await request.json()
    con = db_connect()
    user = con.execute(
        "SELECT id,username,name,role FROM users WHERE username=? AND password=?",
        (data.get("username", ""), data.get("password", "")),
    ).fetchone()
    con.close()
    if not user:
        raise HTTPException(401, "Usuário ou senha inválidos.")
    return dict(user)


@app.get("/api/users")
def list_users():
    con = db_connect()
    rows = con.execute("SELECT id,username,name,role FROM users ORDER BY name").fetchall()
    con.close()
    return [dict(r) for r in rows]


@app.post("/api/uploads")
async def upload_files(files: list[UploadFile] = File(...)):
    saved: list[str] = []
    for file in files:
        suffix = Path(file.filename or "arquivo").suffix.lower()
        safe = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{suffix}"
        path = UPLOAD_DIR / safe
        path.write_bytes(await file.read())
        saved.append(f"/uploads/{safe}")
    return {"paths": saved}


@app.post("/api/import-excel")
async def import_excel(request: Request, file: UploadFile = File(...)):
    user_id = int(request.query_params.get("user_id", "0") or 0)
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Envie um arquivo Excel .xlsx ou .xlsm.")
    content = await file.read()
    temp = DATA_DIR / f"import_{uuid.uuid4().hex}.xlsx"
    temp.write_bytes(content)
    wb = None
    try:
        wb = load_workbook(temp, read_only=True, data_only=True)
        if "Relatorio de Compras" not in wb.sheetnames:
            raise HTTPException(400, "A aba 'Relatorio de Compras' não foi encontrada.")
        ws = wb["Relatorio de Compras"]
        iterator = ws.iter_rows(values_only=True)
        raw_headers = next(iterator, None)
        if not raw_headers:
            raise HTTPException(400, "A planilha está vazia.")
        headers = {normalize_header(v): i for i, v in enumerate(raw_headers)}
        required = ["idcompra", "statuscompra", "statuskanban", "tipo", "produto", "qtdesperada"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise HTTPException(400, "Colunas obrigatórias ausentes: " + ", ".join(missing))

        def cell(row: tuple[Any, ...], name: str, default: Any = None) -> Any:
            idx = headers.get(name)
            return default if idx is None or idx >= len(row) else row[idx]

        con = db_connect()
        if user_id:
            require_role(con, user_id, {"admin"})
        total_rows = matched_rows = cards_created = cards_updated = items_created = items_updated = cards_removed_cd02 = 0
        touched_cards: set[int] = set()
        errors: list[str] = []
        for line_no, row in enumerate(iterator, start=2):
            total_rows += 1
            # O relatório mistura compras finalizadas e futuras. Para o painel operacional,
            # importamos integralmente as compras em trânsito ou já dentro da Operação.
            purchase_status = normalize_text(cell(row, "statuscompra", ""))
            if purchase_status not in {"transito", "operacoes"}:
                continue
            matched_rows += 1
            try:
                purchase_id = clean_id(cell(row, "idcompra"))
                if not purchase_id:
                    raise ValueError("ID Compra vazio")
                existing = con.execute("SELECT * FROM cards WHERE purchase_id=?", (purchase_id,)).fetchone()
                original_type = str(cell(row, "tipo", "") or "").strip()
                purchase_mode = purchase_mode_from_type(original_type)
                if not purchase_mode:
                    raise ValueError(f"Tipo da compra não reconhecido: {original_type or 'vazio'}")
                card_data = (
                    clean_date(cell(row, "datacriacao")),
                    str(cell(row, "fornecedor", "") or "").strip(),
                    original_type,
                    purchase_mode,
                    str(cell(row, "statuscompra", "") or "").strip(),
                    str(cell(row, "destino", "") or "").strip(),
                    clean_date(cell(row, "dataentregaprevista")),
                    clean_int(cell(row, "qtditens")),
                    str(cell(row, "observacoes", "") or "").strip(),
                    str(cell(row, "marca", "") or "").strip(),
                    str(cell(row, "colecao", "") or "").strip(),
                )
                if existing:
                    con.execute(
                        """UPDATE cards SET source_created_date=?,supplier=?,original_type=?,purchase_mode=?,status_compra=?,
                           original_destination=?,forecast_date=?,qtd_itens=?,source_notes=?,brand=?,collection=?,updated_at=?
                           WHERE id=?""",
                        (*card_data, iso_now(), existing["id"]),
                    )
                    card_id = existing["id"]
                    if card_id not in touched_cards:
                        cards_updated += 1
                else:
                    cur = con.execute(
                        """INSERT INTO cards(purchase_id,source_created_date,supplier,original_type,purchase_mode,status_compra,
                           original_destination,forecast_date,qtd_itens,source_notes,brand,collection,current_sector,status,
                           receiving_type,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (purchase_id, *card_data, "RECEBIMENTO", "AGUARDANDO_RECEBIMENTO", "NOVA", iso_now(), iso_now()),
                    )
                    card_id = cur.lastrowid
                    cards_created += 1
                    add_history(con, card_id, "IMPORTACAO", f"Compra {purchase_id} importada diretamente para Recebimento.", user_id or None)
                touched_cards.add(card_id)

                product = str(cell(row, "produto", "") or "").strip()
                sku = str(cell(row, "sku", "") or "").strip()
                source_key = "|".join([
                    purchase_id,
                    clean_id(cell(row, "idlote")),
                    sku or product,
                    str(cell(row, "cor", "") or "").strip(),
                    str(cell(row, "tamanho", "") or "").strip(),
                    str(cell(row, "referencia", "") or "").strip(),
                ])
                source_values = {
                    "status_kanban": cell(row, "statuskanban", ""),
                    "status_purchase": cell(row, "statuscompra", ""),
                    "status_lot": cell(row, "statuslote", ""),
                    "status_quality": cell(row, "statusqualidade", ""),
                    "inspection_phase": cell(row, "faseinspecao", ""),
                    "status_pcp": cell(row, "statuspcp", ""),
                    "status_logistics": cell(row, "statuslogistica", ""),
                }
                item_values = (
                    product,
                    str(cell(row, "referencia", "") or "").strip(),
                    sku,
                    str(cell(row, "grupo", "") or "").strip(),
                    str(cell(row, "colecao", "") or "").strip(),
                    str(cell(row, "marca", "") or "").strip(),
                    str(cell(row, "genero", "") or "").strip(),
                    str(cell(row, "cor", "") or "").strip(),
                    str(cell(row, "tamanho", "") or "").strip(),
                    str(cell(row, "capsula", "") or "").strip(),
                    str(cell(row, "lote", "") or "").strip(),
                    clean_id(cell(row, "idlote")),
                    str(cell(row, "nf", "") or "").strip(),
                    clean_int(cell(row, "qtdesperada")),
                    str(cell(row, "urlfoto", "") or "").strip(),
                    str(cell(row, "statuskanban", "") or "").strip(),
                    source_stage_from_row(source_values),
                    str(source_values["status_purchase"] or "").strip(),
                    str(source_values["status_lot"] or "").strip(),
                    str(source_values["status_quality"] or "").strip(),
                    str(source_values["inspection_phase"] or "").strip(),
                    str(source_values["status_pcp"] or "").strip(),
                    str(cell(row, "costureiro", "") or "").strip(),
                    str(source_values["status_logistics"] or "").strip(),
                    clean_int(cell(row, "qtdrecebida")),
                )
                item = con.execute("SELECT id FROM items WHERE card_id=? AND source_key=?", (card_id, source_key)).fetchone()
                if item:
                    con.execute(
                        """UPDATE items SET product=?,reference=?,sku=?,group_name=?,collection=?,brand=?,gender=?,color=?,
                           size=?,capsule=?,lot=?,lot_id=?,nf=?,expected_qty=?,url_photo=?,status_kanban=?,source_stage=?,
                           source_status_purchase=?,source_status_lot=?,source_status_quality=?,source_inspection_phase=?,
                           source_status_pcp=?,source_seamstress=?,source_status_logistics=?,source_received_qty=? WHERE id=?""",
                        (*item_values, item["id"]),
                    )
                    items_updated += 1
                else:
                    con.execute(
                        """INSERT INTO items(card_id,source_key,product,reference,sku,group_name,collection,brand,gender,
                           color,size,capsule,lot,lot_id,nf,expected_qty,url_photo,status_kanban,source_stage,
                           source_status_purchase,source_status_lot,source_status_quality,source_inspection_phase,
                           source_status_pcp,source_seamstress,source_status_logistics,source_received_qty)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (card_id, source_key, *item_values),
                    )
                    items_created += 1
            except Exception as exc:
                errors.append(f"Linha {line_no}: {exc}")

        for card_id in touched_cards:
            stages = {row["source_stage"] for row in con.execute("SELECT source_stage FROM items WHERE card_id=?", (card_id,)).fetchall()}
            status, summary = card_source_status(stages)
            card = con.execute("SELECT original_destination,current_sector,status FROM cards WHERE id=?", (card_id,)).fetchone()
            if card and is_cd02(card["original_destination"]) and "TRANSITO" not in stages:
                delete_card_with_downstream(con, card_id)
                cards_removed_cd02 += 1
                continue
            sector, status = imported_card_route(stages)
            receiving_type = "RETORNO" if "RETORNO_COSTURA" in stages else "NOVA"
            if card and has_local_workflow(con, card_id, card["current_sector"]):
                con.execute("""UPDATE cards SET source_location_summary=?,source_snapshot_at=?,updated_at=?
                    WHERE id=?""", (summary, iso_now(), iso_now(), card_id))
            else:
                con.execute("""UPDATE cards SET current_sector=?,status=?,receiving_type=?,
                    source_location_summary=?,source_snapshot_at=?,updated_at=? WHERE id=?""",
                    (sector, status, receiving_type, summary, iso_now(), iso_now(), card_id))
            # Só abre um recebimento operacional quando a fonte diz que a mercadoria está em trânsito.
            # As demais posições são exibidas no Recebimento como fotografia fiel do relatório.
            if stages == {"TRANSITO"}:
                ensure_receiving(con, card_id, "NOVA")
        con.execute(
            """INSERT INTO imports(filename,total_rows,matched_rows,cards_created,cards_updated,items_created,items_updated,
               errors,user_id,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (file.filename, total_rows, matched_rows, cards_created, cards_updated, items_created, items_updated,
             "\n".join(errors[:50]), user_id or None, iso_now()),
        )
        con.commit()
        con.close()
        return {
            "total_rows": total_rows,
            "matched_rows": matched_rows,
            "cards_created": cards_created,
            "cards_updated": cards_updated,
            "items_created": items_created,
            "items_updated": items_updated,
            "cards_removed_cd02": cards_removed_cd02,
            "errors": errors[:20],
        }
    finally:
        if wb is not None:
            wb.close()
        temp.unlink(missing_ok=True)


@app.get("/api/imports")
def imports():
    con = db_connect()
    rows = con.execute("SELECT * FROM imports ORDER BY id DESC LIMIT 30").fetchall()
    con.close()
    return [dict(r) for r in rows]


@app.get("/api/dashboard")
def dashboard():
    con = db_connect()
    status_rows = con.execute("SELECT status,COUNT(*) n FROM cards GROUP BY status").fetchall()
    counts = {r["status"]: r["n"] for r in status_rows}
    recent = con.execute(
        """SELECT c.id,c.purchase_id,c.supplier,c.brand,c.purchase_mode,c.current_sector,c.status,c.casulo_current,c.updated_at,
           COALESCE(SUM(i.expected_qty),0) expected_total
           FROM cards c LEFT JOIN items i ON i.card_id=c.id
           GROUP BY c.id ORDER BY c.updated_at DESC LIMIT 12"""
    ).fetchall()
    totals = {
        "cards": con.execute("SELECT COUNT(*) n FROM cards").fetchone()["n"],
        "receiving": con.execute("SELECT COUNT(*) n FROM cards WHERE current_sector='RECEBIMENTO' OR source_snapshot_at IS NOT NULL").fetchone()["n"],
        "pending_10": counts.get("RECEBIMENTO_FISICO_CONCLUIDO_10_PENDENTE", 0),
        "pending_physical": counts.get("SEPARACAO_10_CONCLUIDA_RECEBIMENTO_PENDENTE", 0),
        "awaiting_dispatch": counts.get("AGUARDANDO_DESPACHO_COSTURA", 0),
        "awaiting_return": counts.get("AGUARDANDO_RECEBIMENTO_RETORNO", 0),
        "quality": con.execute("""SELECT COUNT(*) n FROM cards c WHERE c.current_sector='QUALIDADE'
            OR EXISTS (SELECT 1 FROM items i WHERE i.card_id=c.id AND i.source_stage IN
            ('QUALIDADE','QUALIDADE_RETRABALHO','QUALIDADE_REJEITADO'))""").fetchone()["n"],
        "quality_in_progress": counts.get("EM_INSPECAO", 0) + counts.get("INSPECAO_PAUSADA", 0),
        "processing": con.execute("""SELECT COUNT(*) n FROM cards c WHERE c.current_sector='PROCESSAMENTO'
            OR EXISTS (SELECT 1 FROM items i WHERE i.card_id=c.id AND i.source_stage IN
            ('AGUARDANDO_PROCESSAMENTO','PROCESSAMENTO'))""").fetchone()["n"],
        "storage": con.execute("""SELECT COUNT(*) n FROM cards c WHERE c.current_sector='ESTOCAGEM'
            OR EXISTS (SELECT 1 FROM items i WHERE i.card_id=c.id AND i.source_stage='ESTOCAGEM')""").fetchone()["n"],
        "processing_in_progress": counts.get("EM_PROCESSAMENTO", 0) + counts.get("PROCESSAMENTO_PAUSADO", 0),
    }
    con.close()
    return {
        "totals": totals,
        "status_counts": [
            {"status": status, "label": STATUS_LABELS.get(status, status), "value": value}
            for status, value in counts.items()
        ],
        "recent": [dict(r) | {"status_label": STATUS_LABELS.get(r["status"], r["status"])} for r in recent],
    }


@app.get("/api/cards")
def list_cards(scope: str = "receiving", search: str = ""):
    con = db_connect()
    sql = """SELECT c.id,c.purchase_id,c.supplier,c.original_type,c.purchase_mode,c.brand,c.forecast_date,c.current_sector,c.status,
             c.receiving_type,c.quality_destination,c.casulo_current,c.source_location_summary,c.source_snapshot_at,c.updated_at,
             COALESCE(SUM(i.expected_qty),0) expected_total,COUNT(i.id) item_count,
             GROUP_CONCAT(DISTINCT COALESCE(i.product,'') || ' ' || COALESCE(i.reference,'') || ' ' || COALESCE(i.sku,'')) material_search
             FROM cards c LEFT JOIN items i ON i.card_id=c.id WHERE 1=1"""
    params: list[Any] = []
    if scope == "receiving":
        sql += " AND (c.current_sector='RECEBIMENTO' OR c.source_snapshot_at IS NOT NULL)"
    elif scope == "quality":
        sql += " AND (c.current_sector='QUALIDADE' OR EXISTS (SELECT 1 FROM items si WHERE si.card_id=c.id AND si.source_stage IN ('QUALIDADE','QUALIDADE_RETRABALHO','QUALIDADE_REJEITADO')))"
    elif scope == "processing":
        sql += " AND (c.current_sector='PROCESSAMENTO' OR EXISTS (SELECT 1 FROM items si WHERE si.card_id=c.id AND si.source_stage IN ('AGUARDANDO_PROCESSAMENTO','PROCESSAMENTO')))"
    elif scope == "labeling":
        sql += " AND (c.current_sector='ETIQUETAGEM' OR EXISTS (SELECT 1 FROM items si WHERE si.card_id=c.id AND si.source_stage='ETIQUETAGEM'))"
    elif scope == "storage":
        sql += " AND (c.current_sector='ESTOCAGEM' OR EXISTS (SELECT 1 FROM items si WHERE si.card_id=c.id AND si.source_stage='ESTOCAGEM'))"
    elif scope == "outside":
        sql += " AND c.current_sector NOT IN ('RECEBIMENTO','QUALIDADE','PROCESSAMENTO')"
    if search:
        sql += " AND (c.purchase_id LIKE ? OR c.supplier LIKE ? OR c.brand LIKE ? OR c.casulo_current LIKE ? OR i.product LIKE ? OR i.reference LIKE ? OR i.sku LIKE ?)"
        q = f"%{search}%"
        params.extend([q, q, q, q, q, q, q])
    sql += " GROUP BY c.id ORDER BY c.updated_at DESC"
    rows = con.execute(sql, params).fetchall()
    con.close()
    return [dict(r) | {"status_label": STATUS_LABELS.get(r["status"], r["status"])} for r in rows]


@app.get("/api/cards/{card_id}")
def get_card(card_id: int):
    con = db_connect()
    row = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not row:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    card = dict(row)
    card["status_label"] = STATUS_LABELS.get(card["status"], card["status"])
    card["items"] = [dict(r) for r in con.execute("SELECT * FROM items WHERE card_id=? ORDER BY product,color,size,id", (card_id,)).fetchall()]
    card["expected_total"] = sum(item["expected_qty"] for item in card["items"])
    card["receiving"] = current_receiving(con, card_id)
    card["dispatch"] = current_dispatch(con, card_id)
    card["quality"] = quality_card_data(con, card_id)
    card["processing"] = processing_card_data(con, card_id)
    card["labeling"] = downstream_card_data(con, card_id, "ETIQUETAGEM")
    card["storage"] = downstream_card_data(con, card_id, "ESTOCAGEM")
    card["casulo_history"] = [dict(r) for r in con.execute(
        """SELECT ch.*,u.name user_name FROM casulo_history ch JOIN users u ON u.id=ch.user_id
           WHERE ch.card_id=? ORDER BY ch.id DESC""", (card_id,)).fetchall()]
    card["history"] = [dict(r) for r in con.execute(
        """SELECT h.*,u.name user_name FROM history h LEFT JOIN users u ON u.id=h.user_id
           WHERE h.card_id=? ORDER BY h.id DESC""", (card_id,)).fetchall()]
    con.close()
    return card


@app.patch("/api/receivings/{receiving_id}")
async def save_receiving(receiving_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    con = db_connect()
    require_role(con, user_id, {"recebimento"})
    rec = con.execute("SELECT * FROM receivings WHERE id=?", (receiving_id,)).fetchone()
    if not rec:
        con.close()
        raise HTTPException(404, "Recebimento não encontrado.")
    card = con.execute("SELECT * FROM cards WHERE id=?", (rec["card_id"],)).fetchone()
    if card["current_sector"] != "RECEBIMENTO":
        con.close()
        raise HTTPException(400, "O Card não está no Recebimento.")

    def incoming(name: str, current: Any) -> Any:
        return data[name] if name in data else current

    damage = incoming("has_damage", rec["has_damage"])
    if damage is True:
        damage = 1
    elif damage is False:
        damage = 0
    values = (
        incoming("volumes", rec["volumes"]),
        incoming("received_qty", rec["received_qty"]),
        damage,
        incoming("damage_description", rec["damage_description"]),
        incoming("notes", rec["notes"]),
        to_json(incoming("photo_paths", from_json(rec["photo_paths"], []))),
        incoming("ten_percent_actual", rec["ten_percent_actual"]),
        receiving_id,
    )
    con.execute(
        """UPDATE receivings SET volumes=?,received_qty=?,has_damage=?,damage_description=?,notes=?,photo_paths=?,
           ten_percent_actual=? WHERE id=?""",
        values,
    )
    received_qty = incoming("received_qty", rec["received_qty"])
    if received_qty is not None:
        item_count = con.execute(
            "SELECT COUNT(*) n FROM items WHERE card_id=? AND expected_qty>0", (rec["card_id"],)
        ).fetchone()["n"]
        minimum = max(math.ceil(int(received_qty) * 0.10), item_count if card["purchase_mode"] == "GRADE" else 0)
        con.execute("UPDATE receivings SET ten_percent_min=? WHERE id=?", (minimum, receiving_id))
    add_history(con, rec["card_id"], "RECEBIMENTO_SALVO", "Dados do Recebimento atualizados.", user_id)
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/receivings/{receiving_id}/physical-complete")
async def complete_physical(receiving_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    con = db_connect()
    user = require_role(con, user_id, {"recebimento"})
    rec = con.execute("SELECT * FROM receivings WHERE id=?", (receiving_id,)).fetchone()
    if not rec:
        con.close()
        raise HTTPException(404, "Recebimento não encontrado.")
    card = con.execute("SELECT * FROM cards WHERE id=?", (rec["card_id"],)).fetchone()
    if card["current_sector"] != "RECEBIMENTO":
        con.close()
        raise HTTPException(400, "O Card não está no Recebimento.")
    if rec["physical_status"] == "CONCLUIDO":
        con.close()
        raise HTTPException(400, "O recebimento físico já foi concluído.")
    if rec["volumes"] is None or rec["received_qty"] is None:
        con.close()
        raise HTTPException(400, "Informe a quantidade de volumes e a quantidade recebida.")
    if rec["has_damage"] is None:
        con.close()
        raise HTTPException(400, "Informe se existem danos ou avarias.")
    if rec["has_damage"] and not (rec["damage_description"] or "").strip():
        con.close()
        raise HTTPException(400, "Descreva os danos ou avarias.")
    con.execute(
        """UPDATE receivings SET physical_status='CONCLUIDO',physical_completed_by=?,physical_completed_at=?
           WHERE id=?""",
        (user_id, iso_now(), receiving_id),
    )
    add_history(con, rec["card_id"], "RECEBIMENTO_FISICO_CONCLUIDO", f"{user['name']} concluiu o recebimento físico.", user_id)
    next_status = update_new_receiving_flow(con, receiving_id, user_id)
    con.commit()
    con.close()
    return {"ok": True, "next_status": next_status}


@app.post("/api/receivings/{receiving_id}/timer/{action}")
async def sample_timer(receiving_id: int, action: str, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    con = db_connect()
    require_role(con, user_id, {"recebimento"})
    rec = con.execute("SELECT * FROM receivings WHERE id=?", (receiving_id,)).fetchone()
    if not rec:
        con.close()
        raise HTTPException(404, "Recebimento não encontrado.")
    card = con.execute("SELECT * FROM cards WHERE id=?", (rec["card_id"],)).fetchone()
    if card["current_sector"] != "RECEBIMENTO":
        con.close()
        raise HTTPException(400, "O Card não está no Recebimento.")
    if not rec["ten_percent_required"]:
        con.close()
        raise HTTPException(400, "Este recebimento não exige separação dos 10%.")
    if action == "finish":
        rec = con.execute("SELECT * FROM receivings WHERE id=?", (receiving_id,)).fetchone()
        if rec["ten_percent_actual"] is None:
            con.close()
            raise HTTPException(400, "Informe a quantidade efetivamente separada.")
        if rec["ten_percent_actual"] < rec["ten_percent_min"]:
            con.close()
            raise HTTPException(400, f"A quantidade separada deve ser no mínimo {rec['ten_percent_min']} peças.")
    summary = timer_action(con, receiving_id, action, user_id)
    status_map = {"start": "EM_ANDAMENTO", "pause": "PAUSADA", "resume": "EM_ANDAMENTO", "finish": "CONCLUIDA"}
    triage_done = 1 if action == "finish" and rec["receiving_type"] == "NOVA" else rec["triage_completed"]
    con.execute("UPDATE receivings SET ten_percent_status=?,triage_completed=? WHERE id=?",
                (status_map[action],triage_done,receiving_id))
    verbs = {"start": "iniciou", "pause": "pausou", "resume": "retomou", "finish": "concluiu"}
    activity = ("a separação dos 10% e a triagem inicial" if rec["receiving_type"] == "NOVA"
                else "a separação da nova amostra de 10% do retorno CD01")
    add_history(con, rec["card_id"], "SEPARACAO_10", f"Operador {verbs[action]} {activity}.", user_id)
    next_status = None
    if action == "finish":
        next_status = update_new_receiving_flow(con, receiving_id, user_id)
    con.commit()
    con.close()
    return summary | {"next_status": next_status}


@app.post("/api/cards/{card_id}/casulo")
async def update_casulo(card_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    new_casulo = str(data.get("new_casulo") or "").strip()
    note = str(data.get("note") or "").strip()
    if not new_casulo:
        raise HTTPException(400, "Informe o novo Casulo.")
    con = db_connect()
    user = require_role(con, user_id, {"recebimento"})
    card = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not card:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    if card["current_sector"] != "RECEBIMENTO":
        con.close()
        raise HTTPException(400, "O Casulo só pode ser alterado enquanto o Card estiver no Recebimento.")
    old = card["casulo_current"] or ""
    con.execute("UPDATE cards SET casulo_current=?,updated_at=? WHERE id=?", (new_casulo, iso_now(), card_id))
    con.execute(
        """INSERT INTO casulo_history(card_id,old_casulo,new_casulo,note,user_id,created_at)
           VALUES(?,?,?,?,?,?)""",
        (card_id, old, new_casulo, note, user_id, iso_now()),
    )
    add_history(
        con,
        card_id,
        "CASULO_ALTERADO",
        f"{user['name']} alterou o Casulo de '{old or 'não informado'}' para '{new_casulo}'.",
        user_id,
    )
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/cards/{card_id}/dispatch")
async def complete_dispatch(card_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    carrier = str(data.get("carrier") or "").strip()
    seamstress = str(data.get("seamstress_name") or "").strip()
    dispatched_qty = data.get("dispatched_qty")
    volumes = data.get("volumes")
    notes = str(data.get("notes") or "").strip()
    photo_paths = data.get("photo_paths") or []
    if not carrier or not seamstress:
        raise HTTPException(400, "Informe a transportadora e o nome do costureiro.")
    if dispatched_qty in (None, "") or volumes in (None, ""):
        raise HTTPException(400, "Informe a quantidade despachada e os volumes.")
    con = db_connect()
    user = require_role(con, user_id, {"recebimento"})
    card = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not card:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    if card["current_sector"] != "RECEBIMENTO" or card["status"] != "AGUARDANDO_DESPACHO_COSTURA":
        con.close()
        raise HTTPException(400, "O Card não está aguardando despacho no Recebimento.")
    destination = card["quality_destination"]
    if destination not in ("CD01", "CD02"):
        con.close()
        raise HTTPException(400, "O destino da Qualidade não está definido.")
    return_forecast = add_business_days(date.today(), 7).isoformat() if destination == "CD01" else None
    con.execute(
        """INSERT INTO dispatches(card_id,destination,carrier,seamstress_name,dispatched_qty,volumes,photo_paths,
           notes,return_forecast,completed_by,completed_at,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
        (card_id, destination, carrier, seamstress, int(dispatched_qty), int(volumes), to_json(photo_paths),
         notes, return_forecast, user_id, iso_now(), iso_now()),
    )
    if destination == "CD01":
        next_sector, next_status = "RECEBIMENTO", "EM_COSTURA_CD01"
        description = (
            f"{user['name']} despachou a mercadoria para Costura CD01 pela transportadora {carrier}, "
            f"costureiro {seamstress}. Previsão de retorno: {return_forecast}."
        )
    else:
        next_sector, next_status = "FORA_FLUXO", "DESPACHO_CD02"
        description = (
            f"{user['name']} despachou a mercadoria para Costura com destino CD02 pela transportadora {carrier}, "
            f"costureiro {seamstress}. Card encerrado no fluxo interno."
        )
    con.execute(
        "UPDATE cards SET current_sector=?,status=?,updated_at=? WHERE id=?",
        (next_sector, next_status, iso_now(), card_id),
    )
    add_history(con, card_id, "DESPACHO_COSTURA", description, user_id)
    con.commit()
    con.close()
    return {"ok": True, "next_status": next_status}


@app.post("/api/test/cards/{card_id}/quality-return")
async def simulate_quality_return(card_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    destination = str(data.get("destination") or "").strip().upper()
    if destination not in ("CD01", "CD02"):
        raise HTTPException(400, "Selecione CD01 ou CD02.")
    con = db_connect()
    require_role(con, user_id, {"admin"})
    card = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not card:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    if card["current_sector"] != "QUALIDADE":
        con.close()
        raise HTTPException(400, "A simulação exige que o Card esteja encaminhado à Qualidade.")
    con.execute(
        """UPDATE cards SET current_sector='RECEBIMENTO',status='AGUARDANDO_DESPACHO_COSTURA',
           quality_destination=?,receiving_type='NOVA',updated_at=? WHERE id=?""",
        (destination, iso_now(), card_id),
    )
    add_history(
        con,
        card_id,
        "SIMULACAO_QUALIDADE",
        f"Ferramenta de teste: Inspeção 1 concluída com destino {destination}. Card devolvido ao Recebimento para despacho.",
        user_id,
    )
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/test/cards/{card_id}/costura-return")
async def simulate_costura_return(card_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    con = db_connect()
    require_role(con, user_id, {"admin"})
    card = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not card:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    if card["status"] != "EM_COSTURA_CD01":
        con.close()
        raise HTTPException(400, "Somente Cards em Costura CD01 podem retornar.")
    con.execute(
        """UPDATE cards SET current_sector='RECEBIMENTO',status='AGUARDANDO_RECEBIMENTO_RETORNO',
           receiving_type='RETORNO',updated_at=? WHERE id=?""",
        (iso_now(), card_id),
    )
    ensure_receiving(con, card_id, "RETORNO")
    add_history(
        con,
        card_id,
        "SIMULACAO_RETORNO_COSTURA",
        "Ferramenta de teste: retorno CD01 registrado. Card enviado ao Recebimento para conferência e nova tiragem de 10%.",
        user_id,
    )
    con.commit()
    con.close()
    return {"ok": True}


register_quality_routes(app)
register_processing_routes(app)
register_downstream_routes(app)
register_unified_routes(app)

@app.post("/api/test/cards/{card_id}/send-processing")
async def simulate_send_processing(card_id: int, request: Request):
    data = await request.json()
    user_id = int(data.get("user_id") or 0)
    con = db_connect()
    require_role(con, user_id, {"admin"})
    card = con.execute("SELECT * FROM cards WHERE id=?", (card_id,)).fetchone()
    if not card:
        con.close()
        raise HTTPException(404, "Card não encontrado.")
    if card["status"] == "DESPACHO_CD02":
        con.close()
        raise HTTPException(400, "Cards CD02 permanecem fora do fluxo interno.")
    if card["purchase_mode"] not in ("GRADE", "SALDO"):
        con.close()
        raise HTTPException(400, "O tipo da compra não foi reconhecido na importação.")
    con.execute(
        "UPDATE cards SET current_sector='PROCESSAMENTO',status='AGUARDANDO_PROCESSAMENTO',updated_at=? WHERE id=?",
        (iso_now(), card_id),
    )
    add_history(con, card_id, "SIMULACAO_PROCESSAMENTO", "Ferramenta de teste: Card encaminhado diretamente ao Processamento.", user_id)
    con.commit()
    con.close()
    return {"ok": True}

@app.get("/api/history")
def global_history(limit: int = 300):
    con = db_connect()
    rows = con.execute(
        """SELECT h.*,c.purchase_id,u.name user_name FROM history h
           JOIN cards c ON c.id=h.card_id LEFT JOIN users u ON u.id=h.user_id
           ORDER BY h.id DESC LIMIT ?""",
        (max(1, min(limit, 1000)),),
    ).fetchall()
    con.close()
    return [dict(r) for r in rows]
